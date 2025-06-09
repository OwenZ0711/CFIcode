#!/usr/bin/python3
import io
import os.path
import subprocess

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import pymssql
import sys

# from toolmodules.modules_sendemail import *
from toolmodules.modules import *

from toolmodules.modules_sendemail import *
from toolmodules.modules_analyze import *
from toolmodules.cfidataapi.trade_info_loader import get_trade_info, trade_agg_func
from toolmodules.utils.parallelism import Parallelism
from toolmodules.utils.classifier_base import GeneralClassifier
from toolmodules.utils.logger import CfiLogger

logger = CfiLogger.get_logger(__file__)


def pre_trading_day(day):
    conn = pymssql.connect('dbs.cfi:1433', 'DataSupply', 'Cfi888_', 'JYDB', charset="utf8", as_dict=True)
    if conn is None:
        raise Exception('db connected error')

    sql = "DECLARE  @QueryDate DATETIME" \
          "\nSET @QueryDate = '%s'" \
          "\nSELECT TOP 1 convert(varchar(10), TradingDate,112) AS TradingDate" \
          "\nFROM" \
          "\n(" \
          "\n        SELECT DISTINCT TradingDate" \
          "\n        FROM QT_TradingDayNew A" \
          "\n        LEFT JOIN CT_SystemConst B ON A.SecuMarket = B.DM" \
          "\n        WHERE B.DM IN (90, 83)" \
          "\n        AND B.LB=201" \
          "\n        AND A.IfTradingDay=1" \
          "\n        AND A.TradingDate < @QueryDate" \
          "\n        ) AA" \
          "\nORDER BY TradingDate DESC;" % day

    cursor = conn.cursor()
    cursor.execute(sql)
    row = cursor.fetchone()

    if row is None:
        raise Exception('SQL excuted error')

    return row["TradingDate"]


def next_trading_day(day):
    conn = pymssql.connect('dbs.cfi:1433', 'DataSupply', 'Cfi888_', 'JYDB', charset="utf8", as_dict=True)
    if conn is None:
        raise Exception('db connected error')

    sql = "DECLARE @QueryDate DATETIME" \
          "\nSET @QueryDate = '%s'" \
          "\nSELECT TOP 1 convert(varchar(10), TradingDate,112) AS TradingDate" \
          "\nFROM" \
          "\n(" \
          "\n        SELECT DISTINCT TradingDate" \
          "\n        FROM QT_TradingDayNew A" \
          "\n        LEFT JOIN CT_SystemConst B ON A.SecuMarket = B.DM" \
          "\n        WHERE B.DM IN (90, 83)" \
          "\n        AND B.LB=201" \
          "\n        AND A.IfTradingDay=1" \
          "\n        AND A.TradingDate > @QueryDate" \
          "\n        ) AA" \
          "\nORDER BY TradingDate ASC;" % day

    cursor = conn.cursor()
    cursor.execute(sql)
    row = cursor.fetchone()

    if row is None:
        raise Exception('SQL excuted error')

    # conn.close()
    return row["TradingDate"]


def get_trader_cfg(date):
    colo_list = get_colo_list(date)
    account_df_list = []
    for colo in colo_list:
        trader_list_path = f"{PLATFORM_PATH_DICT['z_path']}ProcessedData/PTA/parsed_log/{colo}/{date}/trader-list.cfg"
        if not os.path.exists(trader_list_path):
            account_df_list.append(
                pd.DataFrame(columns=["trade_acc", "remote_colo", "exchange", "colo"])
            )
        else:
            with io.open(trader_list_path, encoding='utf-8') as f:
                logger.info(trader_list_path)
                config = libconf.load(f)
                account_df = pd.DataFrame(
                    [
                        {
                            "trade_acc": t["name"],
                            "remote_colo": t["remote_colo"] if "remote_colo" in t else "",
                            "exchange": t["exchange"] if "exchange" in t else "SZSE"
                        } for t in config["trader"]
                    ]
                )

                if account_df.empty:
                    logger.info(f"{trader_list_path} empty")
                    continue

                account_df["exchange"] = account_df["exchange"].apply(lambda t: "SZ" if t == "SZSE" else t)
                account_df["exchange"] = account_df["exchange"].apply(lambda t: "SH" if t == "SHSE" else t)
                account_df["remote_colo"].replace("", np.nan, inplace=True)

                account_df["colo"] = colo
                account_df_list.append(account_df)

    return pd.concat(account_df_list, names=["trade_acc", "remote_colo", "exchange", "colo"], axis=0)


def get_colo_list(date):
    """
    """
    engine = create_engine(f"mssql+pymssql://datasupply:Cfi888_@dbs.cfi/DataSupply?charset=utf8")
    colo_list = pd.read_sql(f"select distinct [colo] from trading_account where [date] = {date}", con=engine)["colo"].tolist()

    if len(colo_list) == 0:
        colo_list = open(f"{PLATFORM_PATH_DICT['z_path']}webmonitor_data/colo_list_trading.txt", "r").read().strip().split(",")
        colo_list = list(set(colo_list) - set(Temp_Colo_Mechine_Dict['simulist']))

        extra_colo_list = []
        black_list = []
        colo_list = [t for t in colo_list if t not in black_list] + extra_colo_list

    normal_colo_list = [t for t in colo_list if t.count("-") == 2]
    abnormal_colo_list = sorted([t for t in set(colo_list) if t.count("-") != 2])

    colo_list = normal_colo_list
    colo_array = sorted(["-".join(np.array(t.split("-"))[[0,2,1]]) if t.count("-") > 1 else t for t in colo_list])
    colo_list = ["-".join(np.array(str(t).split("-"))[[0,2,1]]) if str(t).count("-") > 1 else t for t in colo_array]
    return colo_list + abnormal_colo_list


def read_trading_account_from_db(date):
    engine = create_engine(f"mssql+pymssql://datasupply:Cfi888_@dbs.cfi/DataSupply?charset=utf8")
    trading_account_df = pd.read_sql(f"select * from trading_account where [date]='{date}'", engine)
    return trading_account_df


def run_cmd(cmd, host=None, port=None, **kwargs):
    run_args = { 
        "check": True,
    }
    run_args.update(kwargs)
    if host:
        cmd_arg = ["ssh", "-q", str(host)]
        if port:
            cmd_arg.extend(["-p", str(int(port))])
        cmd_arg.append(cmd)
        logger.info(f"Running cmd:{cmd_arg}")
        logger.info(f"Running cmd," + " ".join(f"'{_x}'" for _x in cmd_arg))
        return subprocess.run(cmd_arg, **run_args)
    else:
        cmd_arg = cmd 
        if isinstance(cmd, str):
            run_args["shell"] = True
            logger.info(f"Running cmd,{cmd_arg}")
            return subprocess.run(cmd_arg, **run_args)
        else:
            logger.info(f"Running cmd," + " ".join(f"'{_x}'" for _x in cmd_arg))
            return subprocess.run(cmd_arg, **run_args)


class PtaTaskProfile(object):
    def __init__(self):
        # with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "trade_logs_profile.json"),
        #           "r", encoding='utf-8') as f:
        #     print(f.read())
        #     self.trade_log_profile = json.loads(f.read())

        self.trade_log_profile = {
            "cpuinfo": {
                "AMD EPYC 7742 64-Core Processor": 2.25,
                "AMD EPYC 7662 64-Core Processor": 2.0,
                "AMD EPYC 7713 64-Core Processor": 2.0,
                "AMD EPYC 7713P 64-Core Processor": 2.0,
                "AMD EPYC 7453 28-Core Processor": 2.75,
                "AMD EPYC 7H12 64-Core Processor": 2.6,
                "AMD EPYC 7452 32-Core Processor": 2.35,
                "Intel(R) Xeon(R) Gold 6258R CPU @ 2.70GHz": 2.7,
                "Intel(R) Xeon(R) Platinum 8268 CPU @ 2.90GHz": 2.9,
                "Intel(R) Xeon(R) Gold 6254 CPU @ 3.10GHz": 3.1,
                "Intel(R) Xeon(R) Gold 6248R CPU @ 3.00GHz": 3.0,
                "Intel(R) Xeon(R) Gold 6354 CPU @ 3.00GHz": 3.0,
                "Intel(R) Xeon(R) Gold 6240 CPU @ 2.60GHz": 2.60,
                "Intel(R) Xeon(R) Gold 6240R CPU @ 2.40GHz": 2.40,
                "AMD EPYC 7763 64-Core Processor": 2.45,
                "AMD EPYC 7352 24-Core Processor": 2.3,
            },
            "jumper_info": {
                "username": "jumper",
                "host": "139.196.57.140",
                "port": 22,
                "parsed_log": "/home/jumper/trade_log/data"
            },
            "machines": {
                "ct-sh-13": {
                    "jumper_info": {
                        "username": "jumper",
                        "host": "139.196.57.140",
                        "port": 22,
                        "parsed_log": "/home/jumper/trade_log/data"
                    }
                }
            },
            "SpecialCare": {
                "types": {
                    "T0_SwapProduction": T0_SwapProductionList,
                    "Production_OwnList": Production_OwnList,
                },
                "DisplayColumns": [
                    "production", "type", "colo", "stid", "总订单", "插单延迟(SZ_broker median)", "插单延迟API(SZ_broker median)",
                    "插单延迟(SZ_exch median)", "插单延迟API(SZ_exch median)", "插单延迟(SZ_exch_buy median)",
                    "插单延迟API(SZ_exch_buy median)",
                    "插单延迟(SZ_exch_sell median)", "插单延迟API(SZ_exch_sell median)",
                    "production", "type", "colo", "成交延迟API(SZ)", "插单延迟(SH_broker median)",
                    "插单延迟API(SH_broker median)", "插单延迟(SH_exch median)", "插单延迟API(SH_exch median)", "成交延迟API(SH)",
                    "production", "type", "colo", "撤单股数比(SZ)", "撤单金额比(SZ)", "撤单股数比(SH)", "撤单金额比(SH)", "撤单股数比",
                    "撤单金额比", "SH_tick2trade(median)", "SZ_tick2trade(median)"
                ]
            },
            "exception": {
                "SZ_tick2trade(median)": {
                    "upper": 0.18,
                    "drift_ratio": 1
                },
                "SH_tick2trade(median)": {
                    "upper": 1.8,
                    "drift_ratio": 2
                },
                "撤单金额比(SZ)": {
                    "drift_ratio": 0.6
                },
                "撤单金额比(SH)": {
                    "drift_ratio": 0.6
                }
            },
            "worker_download": [
                "ct-sz-11"
            ],
            "worker_download.old": [
                "gj-sh-4",
                "gj-sz-4"
            ],
            "trader_err_code_map": {
                "trader1": {
                    "27": "资金不足",
                    "28": "股份不足",
                    "127": "该账户未登录",
                    "1182": "可用数量小于最低卖出数量必须全部卖出",
                    "1183": "卖出零股数量与可用数量的零股数量不相同",
                    "1187": "卖出资金小于费用，不允许交易",
                    "1002": "废单",
                    "300028": "账户持有股份不足"
                },
                "trader2": {
                    "507": "未通过白名单检查",
                    "502": "防对敲控制",
                    "509": "价格偏离度控制",
                    "523": "拉台打压风控",
                    "517": "证券持仓数量控制"
                },
                "trader3": {
                    "251005": "股份不足",
                    "260200": "可用资金不足"
                },
                "trader4": {
                    "11100000": "订单被交易所拒绝"
                },
                "traderDMA": {
                    "111049": "仓位不足"
                },
                "traderHuaXin": {
                    "-2": "超过流速权",
                    "322": "申报量超出单笔申报限制",
                    "360": "资金不足",
                    "361": "数量不足",
                    "403": "投资者代码与账户不一致",
                    "421": "超出买入上限",
                    "433": "华鑫Dma 大账户内部风控-自成交"
                }
            },
            "err_code_map": {
                "csc": "trader1",
                "gj": "trader1",
                "gx": "trader1",
                "gf": "trader1",
                "htsc": "trader1",
                "htsc-sz-5": "trader2",
                "ct": "trader3",
                "cicc": "trader1",
                "zt": "trader4",
                "citic": "trader1",
                "cfipasz1": "trader1",
                "gx-sz-5": "traderDMA",
                "cf": "traderHuaXin"
            },
            "notify": {
                "email": {
                    "smtp_server": "smtp.exmail.qq.com",
                    "account": "andrewyan@centuryfrontier.com",
                    "passwd": "gZ3sMDxLRgN5cFki",
                    "sendto": [
                        "luom@centuryfrontier.com",
                        "1026688756@qq.com",
                        "zhzhou@centuryfrontier.com",
                        "hyanak@centuryfrontier.com",
                        "24085778@qq.com",
                        "471743716@qq.com",
                        "andrewyan@centuryfrontier.com",
                        "sun@centuryfrontier.com",
                        "wenzan@centuryfrontier.com",
                        "rczhao@centuryfrontier.com",
                        "ritchguo@centuryfrontier.com",
                        "yang@centuryfrontier.com",
                        "duanlian@centuryfrontier.com"
                    ],
                    "sendto1": [
                        "duanlian@centuryfrontier.com",
                        "andrewyan@centuryfrontier.com"
                    ]
                }
            },
            "err_notify": [
                "luom@centuryfrontier.com",
                "zhzhou@centuryfrontier.com",
                "hyanak@centuryfrontier.com",
                "471743716@qq.com",
                "24085778@qq.com",
                "rczhao@centuryfrontier.com",
                "andrewyan@centuryfrontier.com",
                "matt@centuryfrontier.com",
                "ritchguo@centuryfrontier.com"
            ],
            "local_path": {
                "PTA_ROOT": "/data",
                "report": "/data/PTA/report",
                "parsed_log": "/data/PTA/parsed_log",
                "alpha_trade_extract": "/data/PTA/ExtractRawData/alpha_trade",
                "alpha_trade_proceed": "/data/PTA/ProcessedData/alpha_trade",
                "chart": "/data/PTA/chart"
            },
            "nas_path": {
                "report": f"{PLATFORM_PATH_DICT['z_path']}ProcessedData/PTA/report",
                "parsed_log": f"{PLATFORM_PATH_DICT['z_path']}ProcessedData/PTA/parsed_log",
                "parsed_log.old": f"{PLATFORM_PATH_DICT['t_path']}ProcessedData/PTA/parsed_log",
                "chart": f"{PLATFORM_PATH_DICT['z_path']}ProcessedData/PTA/chart"
            },
            "nas_raw_dir": f"{PLATFORM_PATH_DICT['t_path']}RawData/PTA"
        }

    def special_care_production(self):
        return self.trade_log_profile["SpecialCare"]["types"]

    def special_care_display_columns(self):
        return self.trade_log_profile["SpecialCare"]["DisplayColumns"]

    def local_report_dir(self):
        return self.trade_log_profile["local_path"]["report"]

    def local_parsed_dir(self):
        return self.trade_log_profile["local_path"]["parsed_log"]

    def local_chart_dir(self):
        return self.trade_log_profile["local_path"]["chart"]

    def nas_report_dir(self):
        return self.trade_log_profile["nas_path"]["report"]

    def nas_parsed_dir(self):
        return self.trade_log_profile["nas_path"]["parsed_log"]

    def nas_parsed_dir_old(self):
        return self.trade_log_profile["nas_path"]["parsed_log.old"]

    def email_server(self):
        return self.trade_log_profile["notify"]["email"]["smtp_server"]

    def email_account(self):
        return self.trade_log_profile["notify"]["email"]["account"]

    def email_passwd(self):
        return self.trade_log_profile["notify"]["email"]["passwd"]

    def email_send_to(self):
        return self.trade_log_profile["notify"]["email"]["sendto"]

    def err_email_send_to(self):
        return self.trade_log_profile["err_notify"]

    def machines_keys(self):
        return list(self.trade_log_profile["machines"].keys())

    def cpu_name(self, colo_name, date):
        machine_info_file_path = os.path.join(self.local_parsed_dir(), colo_name, date, "machine_info.txt")
        if not os.path.exists(machine_info_file_path):
            machine_info_file_path = os.path.join(self.nas_parsed_dir(), colo_name, date, "machine_info.txt")

        # cmd = f"cat {machine_info_file_path} | awk -F '[:：]' '/^型号名称：|^Model name:/ {{print $2$4}}' | sed 's/^ *//g'"
        # ret = subprocess.run(cmd, stdout=subprocess.PIPE, shell=True)
        # cpu_name = str(ret.stdout, encoding="utf-8").strip()

        machine_info_file_path = machine_info_file_path.replace('\\', '/')
        cmd = f""""C:/Program Files/Git/usr/bin/cat.exe" {machine_info_file_path} | "C:/Program Files/Git/usr/bin/awk.exe" -F "[:：]" "/^Model name:|^型号名称：/ {{print $2$4}}" | "C:/Program Files/Git/usr/bin/sed.exe" "s/^ *//g" """
        print(cmd)
        status, cpu_name = subprocess.getstatusoutput(cmd)
        print(cpu_name)
        return cpu_name

    def machine_cpu_freq(self, colo_name, date):
        machine_info_file_path = os.path.join(self.local_parsed_dir(), colo_name, date, "machine_info.txt")
        if not os.path.exists(machine_info_file_path):
            machine_info_file_path = os.path.join(self.nas_parsed_dir(), colo_name, date, "machine_info.txt")

        machine_info_file_path = machine_info_file_path.replace('\\', '/')
        cmd = f""""C:/Program Files/Git/usr/bin/awk.exe" '/Refined TSC clocksource calibration:/ {{ if ($9 == "MHz") print $8}}' {machine_info_file_path}"""
        print(cmd)
        status, tsc_str = subprocess.getstatusoutput(cmd)
        print(status, tsc_str)
        # cmd = f"""awk '/Refined TSC clocksource calibration:/ {{ if ($9 == "MHz") print $8}}' {machine_info_file_path}"""
        # tsc_ret = subprocess.run(cmd, stdout=subprocess.PIPE, shell=True)
        # tsc_str = str(tsc_ret.stdout, encoding="utf-8").strip()
        if tsc_str != "":
            return float(tsc_str) * 1e3

        logger.info(f"{colo_name} {date} cpuinfo not found")
        cpu_name = self.cpu_name(colo_name, date)

        if cpu_name not in self.trade_log_profile["cpuinfo"]:
            logger.info(colo_name, cpu_name)

        return self.trade_log_profile["cpuinfo"][cpu_name] * 1e6

    def jumper_info(self, hostname):
        if hostname in self.trade_log_profile["machines"] and "jumper_info" in self.trade_log_profile["machines"][hostname]:
            return self.trade_log_profile["machines"][hostname]["jumper_info"]
        else:
            return self.trade_log_profile["jumper_info"]

    def machine_user(self, hostname):
        return self.jumper_info(hostname)["username"]

    def machine_ip(self, hostname):
        return self.jumper_info(hostname)["host"]

    def machine_port(self, hostname):
        return self.jumper_info(hostname)["port"]

    def machine_parsed_log(self, hostname):
        return self.jumper_info(hostname)["parsed_log"]

    def colo_worker_download(self):
        return self.trade_log_profile["worker_download"]

    def nas_raw_dir(self):
        return self.trade_log_profile["nas_raw_dir"]

    def exception(self):
        exception_dict = self.trade_log_profile["exception"]
        return exception_dict 

    def pta_local_root(self):
        return self.trade_log_profile["PTA_ROOT"]

    def download_dir(self, category):
        return os.path.join(self.pta_local_root(), "RawData", category)

    def extract_dir(self, category):
        return os.path.join(self.pta_local_root(), "ExtractRawData", category)

    def intermediate_dir(self, category):
        return os.path.join(self.pta_local_root(), "ProcessedData", category)

    def clean_up_remote_dag(self):
        def _handle(**context):
            task_date = context["tomorrow_ds_nodash"]
            today = datetime.datetime.now().strftime("%Y%m%d")
            machines = self.machines_keys()
            cmd_list = []
            for m_key in machines:
                host_ip = self.machine_ip(m_key)
                user = self.machine_user(m_key)
                port = self.machine_port(m_key)
                data_dir = self.machine_parsed_log(m_key)
                if today == task_date:
                    cmd = f"rm -rf {data_dir}"
                else:
                    cmd = f"find {data_dir} -mindepth 3 -type d | grep -v {today} | xargs rm -rf"

                cmd_list.append((cmd, f"{user}@{host_ip}", port))

            for cmd_para in set(cmd_list):
                run_cmd(cmd_para[0], cmd_para[1], cmd_para[2])

        return _handle


pta_task_profile = PtaTaskProfile()


def load_alpha_order_md(op_date, production):
    order_md_dir = f"/data2/andrew/market_impact/data/{op_date}/arranged/alpha_order_md/{op_date}/{production}"
    if not os.path.exists(order_md_dir):
        return pd.DataFrame(columns=["insid", "wf-orderid"])

    para_list = [(os.path.join(order_md_dir, t), ) for t in os.listdir(order_md_dir)]
    pl = Parallelism()
    pl.run(pd.read_csv, para_list)
    order_md_df = pd.concat(pl.get_results(), axis=0)
    return order_md_df 


def freq_analysis(op_date, output_dir=None, month_end=True):
    if output_dir is None:
        output_dir = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{op_date}/OrderFreq/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    morning_timeindex_idx = pd.timedelta_range(start="09:30:00", end="11:30:00", freq="1s")
    after_timeindex_idx = pd.timedelta_range(start="13:00:00", end="15:00:00", freq="1s")
    timeindex_idx = morning_timeindex_idx.append(after_timeindex_idx)

    time_dict = {
        str(t)[7:15]: t for t in timeindex_idx 
    }

    classifier = GeneralClassifier(time_dict)
    trade_cfg_df = read_trading_account_from_db(op_date)

    colo_list = get_colo_list(op_date)
    # colo_list = ['cf-sz-1']
    trader_list_cfg_df = get_trader_cfg(op_date)
    cfg_df = pd.merge(trader_list_cfg_df, trade_cfg_df[["colo",  "stid", "trade_acc", "production"]], on=["colo", "trade_acc"], how="inner")
    cfg_df = cfg_df[cfg_df['colo'].isin(colo_list)]

    pta_task_profile = PtaTaskProfile()
    colo_2_machine_cpu_freq = {
        colo: pta_task_profile.machine_cpu_freq(colo, op_date) for colo in cfg_df['colo'].unique()}
    cfg_df['cpu_freq'] = cfg_df['colo'].apply(lambda x: colo_2_machine_cpu_freq[x])

    order_list = []
    for colo in colo_list:
        nas_parsed_dir = pta_task_profile.nas_parsed_dir()
        csv_path = os.path.join(nas_parsed_dir, colo, op_date, f"{colo}-worker-{op_date}.csv")
        if not os.path.exists(csv_path):
            nas_parsed_dir_old = pta_task_profile.nas_parsed_dir_old()
            csv_path = os.path.join(nas_parsed_dir_old, colo, op_date, f"{colo}-worker-{op_date}.csv")
        order_df = pd.read_csv(csv_path)
        if order_df.empty:
            continue
        order_df = order_df.sort_values('md_localtime', ascending=True)

        md_start_time = pd.Timedelta(order_df['md_exchtime'].iloc[0])
        md_end_time = pd.Timedelta(order_df['md_exchtime'].iloc[-1])
        md_local_start_time = order_df['md_localtime'].iloc[0]
        md_local_end_time = order_df['md_localtime'].iloc[-1]

        order_df['cancel_time_md'] = \
            (md_end_time - md_start_time).value / (md_local_end_time - md_local_start_time) * \
            (order_df['cancel_time'] - md_local_start_time) + md_start_time.value

        print(colo, op_date, colo_list.index(colo))
        order_df["colo"] = colo
        order_list.append(order_df)

    order_df = pd.concat(order_list, axis=0)
    order_df = order_df[order_df["send_err_time"].isnull()]
    order_df = pd.merge(order_df, cfg_df[
        ["colo", "stid", "production", "remote_colo", "exchange", "cpu_freq"]], on=["colo", "stid"], how="left")
    order_df["exch"] = order_df["insid"].apply(lambda t: t[:2])
    colo_filter_mask = order_df["remote_colo"].isnull() | (order_df["exchange"] == order_df["exch"])
    order_df = order_df[colo_filter_mask]

    order_df.rename({"orderid": "wf-orderid"}, inplace=True, axis=1)
    order_df['cancel_wait_time'] = (order_df['cancel_time'] - order_df['trader_sent']) / order_df['cpu_freq'] / 1000
    order_df['cancel_in_1s'] = (order_df['cancel_wait_time'] < 1).astype('int')
    # bhprint(order_df)
    # r_df = pd.read_excel("report.xls")
    # r_df.rename({"内部代码": "production"}, axis=1, inplace=True)
    # order_df = pd.merge(r_df, order_df, on="production", how="inner")

    # order_list = []
    # cnt = 0
    # for p, p_order_df in order_df.groupby("production"):
    #     cnt = cnt + 1
    #     print(p, cnt)
    #     alpha_order_md_df = load_alpha_order_md(op_date, p)
    #     alpha_order_md_df.drop_duplicates(["insid", "wf-orderid"], inplace=True)
    #     o_df = pd.merge(p_order_df, alpha_order_md_df[["insid", "wf-orderid"]], on=["insid", "wf-orderid"], how="inner")
    #     order_list.append(o_df)
    # order_df = pd.concat(order_list, axis=0)
    # 上面是读入数据
    
    order_df.reset_index(inplace=True, drop=True)
    order_df["md_exchtime"] = pd.to_timedelta(order_df["md_exc htime"])
    order_df["flag"] = classifier.classify(order_df["md_exchtime"])

    order_df["md_exchtime_shift_500"] = order_df["md_exchtime"] + pd.Timedelta(milliseconds=500)
    order_df["flag_shift"] = classifier.classify(order_df["md_exchtime_shift_500"])
    order_df["send_turnover"] = order_df['volume'] * order_df['price']

    order_df_cancel = order_df[~order_df["cancel_time"].isnull()].copy(deep=True)
    order_df_cancel["cancel_time_md"] = pd.to_timedelta(order_df_cancel["cancel_time_md"])
    order_df_cancel["flag"] = classifier.classify(order_df_cancel["cancel_time_md"])

    order_df_cancel["cancel_time_md_shift_500"] = order_df_cancel["cancel_time_md"] + pd.Timedelta(milliseconds=500)
    order_df_cancel["flag_shift"] = classifier.classify(order_df_cancel["cancel_time_md_shift_500"])

    # bhprint(order_df)
    order_freq_df = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "code": insid,
                "direct": direct,
                "order_cnt": len(flag_df),
                 "send_turnover": flag_df['send_turnover'].sum(),
                "trade_turnover": flag_df['trade_turnover'].sum(),
                'colo': flag_df['colo'].iloc[0]
            } for (p, flag, exch, insid, direct), flag_df in order_df.groupby(["production", "flag", "exch", "insid", "direction"])
        ]).set_index(["production", "flag", "exch", "code", "direct"])
    
    order_freq_df_cancel = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "code": insid,
                "direct": direct,
                "cancel_cnt": len(flag_df),
                "cancel_num_in_1s": flag_df["cancel_in_1s"].sum(),
                "cancel_diff_time_min": np.nanmin(flag_df["cancel_wait_time"]),
                "cancel_diff_time_max": np.nanmax(flag_df["cancel_wait_time"]),
                "cancel_diff_time_mean": np.nanmean(flag_df["cancel_wait_time"]),
                "cancel_diff_time_median": np.nanmedian(flag_df["cancel_wait_time"]),
            } for (p, flag, exch, insid, direct), flag_df in order_df_cancel.groupby(["production", "flag", "exch", "insid", "direction"])
        ]).set_index(["production", "flag", "exch", "code", "direct"])

    order_freq_df_shift = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "code": insid,
                "direct": direct,
                "order_cnt_shift": len(flag_df),
            } for (p, flag, exch, insid, direct), flag_df in order_df.groupby(["production", "flag_shift", "exch", "insid", "direction"])
        ]).set_index(["production", "flag", "exch", "code", "direct"])

    order_freq_df_cancel_shift = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "code": insid,
                "direct": direct,
                "cancel_cnt_shift": len(flag_df),
            } for (p, flag, exch, insid, direct), flag_df in order_df_cancel.groupby(["production", "flag_shift", "exch", "insid", "direction"])
        ]).set_index(["production", "flag", "exch", "code", "direct"])

    order_freq_df_new = pd.concat([order_freq_df, order_freq_df_cancel, order_freq_df_shift, order_freq_df_cancel_shift], axis=1).fillna(0).reset_index()

    # order_freq_df_new.to_csv(f"{output_dir}{op_date}_order_freq_1s_details_new.csv", index=False)
    order_freq_df_new.to_parquet(f"{output_dir}{op_date}_order_freq_1s_details_new.parquet")

    order_freq_df = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                'colo': flag_df['colo'].iloc[0],
                "order_cnt": flag_df['order_cnt'].sum(),
                "order_cnt_shift": flag_df['order_cnt_shift'].sum(),
                "cancel_cnt": flag_df['cancel_cnt'].sum(),
                "cancel_cnt_shift": flag_df['cancel_cnt_shift'].sum(),
                "send_turnover": flag_df['send_turnover'].sum(),
                "trade_turnover": flag_df['trade_turnover'].sum(),
                "cancel_num_in_1s": flag_df["cancel_num_in_1s"].sum(),
                "cancel_diff_time_min": np.nanmin(flag_df["cancel_wait_time"].replace(0, np.nan)),
                "cancel_diff_time_max": np.nanmax(flag_df["cancel_wait_time"].replace(0, np.nan)),
                "cancel_diff_time_mean": np.nanmean(flag_df["cancel_wait_time"].replace(0, np.nan)),
                "cancel_diff_time_median": np.nanmedian(flag_df["cancel_wait_time"].replace(0, np.nan)),
            } for (p, flag, exch), flag_df in order_freq_df_new.groupby(["production", "flag", "exch"])
        ])
    order_freq_df.to_parquet(f"{output_dir}{op_date}_order_freq_1s_details.parquet")

    order_freq_df_code = pd.DataFrame(
        [
            {
                "production": p,
                "code": code,
                "exch": exch,
                'colo': flag_df['colo'].iloc[0],
                
                "order_cnt": flag_df['order_cnt'].sum(),
                "cancel_cnt": flag_df['cancel_cnt'].sum(),
                
                "send_turnover": flag_df['send_turnover'].sum(),
                "trade_turnover": flag_df['trade_turnover'].sum(),
                "cancel_num_in_1s": flag_df["cancel_num_in_1s"].sum(),
                "cancel_diff_time_min": np.nanmin(flag_df["cancel_wait_time"].replace(0, np.nan)),
                "cancel_diff_time_max": np.nanmax(flag_df["cancel_wait_time"].replace(0, np.nan)),
                "cancel_diff_time_mean": np.nanmean(flag_df["cancel_wait_time"].replace(0, np.nan)),
                "cancel_diff_time_median": np.nanmedian(flag_df["cancel_wait_time"].replace(0, np.nan)),
            } for (p, code, exch), flag_df in order_freq_df_new.groupby(["production", "code", "exch"])
        ])
    order_freq_df_code.to_parquet(f"{output_dir}{op_date}_order_freq_1s_code_details.parquet")

    summary_conlist = []
    for (p, exch), p_df in order_freq_df.groupby(["production", "exch"]):
        p_df = p_df.set_index('flag')
        summary_conlist.append({
            "production": p,
            "exch": exch,
            'colo': p_df['colo'].iloc[0],
            "order_time_max": p_df['order_cnt'].idxmax(),
            "cancel_time_max": p_df['cancel_cnt'].idxmax(),
            "order_and_cancel_time_max": (p_df["order_cnt"] + p_df["cancel_cnt"]).idxmax(),

            "order_freq": np.nanmax(p_df["order_cnt"]),
            "cancel_freq": np.nanmax(p_df["cancel_cnt"]),
            "order_and_cancel_freq": np.nanmax(p_df["order_cnt"] + p_df["cancel_cnt"]),
            "order_freq_shift": np.nanmax(p_df["order_cnt_shift"]),
            "cancel_freq_shift": np.nanmax(p_df["cancel_cnt_shift"]),
            "order_and_cancel_freq_shift": np.nanmax(p_df["order_cnt_shift"] + p_df["cancel_cnt_shift"]),

            "order_daily_sum": np.nansum(p_df["order_cnt"]),
            "cancel_daily_sum": np.nansum(p_df["cancel_cnt"]),
            "order_and_cancel_daily_sum": np.nansum(p_df["order_cnt"] + p_df["cancel_cnt"]),

            'turnover': p_df['trade_turnover'].sum(),
            "cancel_num_in_1s": np.nansum(p_df["cancel_num_in_1s"]),
            "cancel_diff_time_min": np.nanmin(p_df["cancel_diff_time_min"].replace(0, np.nan)),
            "cancel_diff_time_max": np.nanmax(p_df["cancel_diff_time_max"].replace(0, np.nan)),
            "cancel_diff_time_mean": np.nanmean(p_df["cancel_diff_time_mean"].replace(0, np.nan)),
            "cancel_diff_time_median": np.nanmedian(p_df["cancel_diff_time_median"].replace(0, np.nan)),
        })

    summary_conlist_code = []
    for (p, exch), p_df in order_freq_df_code.groupby(["production", "exch"]):
        p_df = p_df.set_index('code')
        order_daily_sum = np.nansum(p_df["order_cnt"])

        p_df['cancel_ratio'] = np.round(p_df['cancel_daily_sum'] / order_daily_sum * 100, 1)
        p_df = p_df.sort_values('cancel_ratio', ascending=False).head(5)

        code_top = ",".join(map(str, p_df.index.to_list()))
        cr_code_top = ",".join(p_df['cancel_ratio'].astype('str').to_list())

        summary_conlist_code.append({
            "production": p,
            "exch": exch,
            'colo': p_df['colo'].iloc[0],
            'TpCode': code_top,
            'CrTpCode': cr_code_top
        })
    order_freq_summary_df = pd.DataFrame(summary_conlist).set_index(['production', 'exch', 'colo'])
    order_freq_summary_df_new = pd.DataFrame(summary_conlist_code).set_index(['production', 'exch', 'colo'])
    order_freq_summary_df = pd.concat([order_freq_summary_df, order_freq_summary_df_new], axis=1).reset_index().sort_values(['production', 'exch'])

    order_freq_summary_df['order_and_cancel_freq_new'] = np.maximum(
        order_freq_summary_df['order_and_cancel_freq'], order_freq_summary_df['order_and_cancel_freq_shift'])
    order_freq_summary_df['cancel_ratio'] = order_freq_summary_df['cancel_daily_sum'] / order_freq_summary_df['order_daily_sum']
    order_freq_summary_df.sort_values("order_and_cancel_freq_new", ascending=False, inplace=True)
    order_freq_summary_df.to_csv(f"{output_dir}{op_date}_order_freq_1s.csv", index=False)
    freq_analysis_zzzq1(op_date)

    report_format = {
        'freq': {
            'SH': {
                (0, 100): '100笔以下',
                (100, 300): '100笔至299笔',
                # (-1, 0): '100笔以下',
                # (0, 300): '100笔至300笔',
                (300, 500): '300笔至499笔',
                (500, 1e9): '500笔及以上',
            },
            'SZ': {
                (0, 100): '0-100',
                (100, 300): '100-300',
                # (-1, 0): '0-100',
                # (0, 300): '100-300',
                (300, 500): '300-500',
                (500, 1e9): '500',
            }
        },
        'order_num': {
            'SH': {
                (0, 10000): '10000笔以下',
                (10000, 15000): '10000笔至14999笔',
                (15000, 20000): '15000笔至19999笔',
                (20000, 25000): '20000笔至24999笔',
                (25000, 1e9): '25000笔及以上',
            },
            'SZ': {
                (0, 10000): '0-10000',
                (10000, 15000): '10000-15000',
                (15000, 20000): '15000-20000',
                (20000, 25000): '20000-25000',
                (25000, 1e9): '25000',
            }
        }
    }

    def format_report_infor(vari, exch, value):
        for freq in report_format[vari][exch].keys():
            if freq[0] <= value < freq[1]:
                return report_format[vari][exch][freq]
        return report_format[vari][exch][freq]

    order_freq_summary_df['Freq'] = order_freq_summary_df.apply(
        lambda row: format_report_infor('freq', row['exch'], row['order_and_cancel_freq_new']), axis=1)
    order_freq_summary_df['OrderNum'] = order_freq_summary_df.apply(
        lambda row: format_report_infor('order_num', row['exch'], row['order_and_cancel_daily_sum']), axis=1)
    order_freq_summary_df[['production', 'exch', 'Freq', 'OrderNum']].to_csv(
        f"{output_dir}{op_date}_order_freq_summary.csv", index=False, encoding='GBK')

    output_dir_w = f'{PLATFORM_PATH_DICT["y_path"]}TradingDailySummary/'
    order_freq_summary_df[['production', 'exch', 'Freq', 'OrderNum']].to_csv(
        f"{output_dir_w}{op_date}_order_freq_summary.csv", index=False, encoding='GBK')

    monthly_date_list = get_monthly_trading_day(op_date)

    if (op_date == monthly_date_list[-1]) and month_end:
        conlist = []
        for date in monthly_date_list:
            df_order_freq = pd.read_csv(
                f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/'
                f'{date}/OrderFreq/{date}_order_freq_1s.csv')

            conlist.append(df_order_freq)

        df_order_freq = pd.concat(conlist, axis=0).groupby(['production', 'exch']).agg(
            {'order_and_cancel_freq_new': 'max', 'order_and_cancel_daily_sum': 'max'}).reset_index()
        df_order_freq['Freq'] = df_order_freq.apply(
            lambda row: format_report_infor('freq', row['exch'], row['order_and_cancel_freq_new']), axis=1)
        df_order_freq['OrderNum'] = df_order_freq.apply(
            lambda row: format_report_infor('order_num', row['exch'], row['order_and_cancel_daily_sum']), axis=1)
        df_order_freq[['production', 'exch', 'Freq', 'OrderNum']].to_csv(
            f"{output_dir_w}{op_date}_order_freq_summary_monthly_max.csv", index=False, encoding='GBK')
        df_order_freq[['production', 'exch', 'Freq', 'OrderNum']].to_csv(
            f"{output_dir}{op_date}_order_freq_summary_monthly_max.csv", index=False, encoding='GBK')
        order_freq_analysis_monthly(monthly_date_list[0], op_date)


def freq_analysis_zzzq1(curdate):
    format_date = curdate[:4] + '-' + curdate[4:6] + '-' + curdate[6:]

    path_zip = f'{TRUSTEESHIP_PATH}zzzq1/109126008616_Orders{format_date}.zip'
    path_order_data = f'{TRUSTEESHIP_PATH}zzzq1/109126008616_Orders{format_date}/'\
                      f'109126008616_Orders{format_date}.csv'

    if not os.path.exists(path_zip):
        return None

    print(path_zip)
    if not os.path.exists(path_order_data):
        try:
            with zipfile.ZipFile(path_zip, 'r') as zip_ref:
                # 解压 ZIP 文件中的所有文件
                zip_ref.extractall(f'{TRUSTEESHIP_PATH}zzzq1/109126008616_Orders{format_date}')
        except:
            print(path_zip, '已经损坏！')
            return None

    if os.path.getsize(path_order_data) == 0:
        return None

    df_order = pd.read_csv(path_order_data)
    df_order['Exchange'] = df_order['证券代码'].apply(
        lambda x: expend_market(expand_stockcode(x), suf_num=2).split('.')[1])

    df_order['委托时间'] = df_order['委托时间'].apply(lambda x: x.strip())
    df_order['撤销时间'] = df_order['撤销时间'].apply(lambda x: x.strip())
    df_order = df_order[df_order['报单状态'] != '已拒绝']
    df_order_cancel = df_order[(df_order['报单状态'] != '全部成交')]
    # df_order_cancel = df_order[(df_order['报单状态'] != '全部成交') &
    #                            (~df_order['撤销时间'].isnull())]

    order_freq_df = pd.DataFrame([{
        "flag": flag.split()[1] if len(flag.split()) == 2 else 'nan',
        "exch": exch,
        "order_cnt": len(flag_df),
        } for (exch, flag), flag_df in df_order.groupby(["Exchange", "委托时间"])])

    if not df_order_cancel.empty:
        order_freq_df_cancel = pd.DataFrame([{
            "flag": flag.split()[1] if len(flag.split()) == 2 else 'nan',
            "exch": exch,
            "cancel_cnt": len(flag_df),
            } for (exch, flag), flag_df in df_order_cancel.groupby(["Exchange", "撤销时间"])])
    else:
        order_freq_df_cancel = pd.DataFrame(columns=['flag', 'exch', 'cancel_cnt'])
    
    order_freq_df = pd.concat([
        order_freq_df.set_index(['flag', 'exch']),
        order_freq_df_cancel.set_index(['flag', 'exch'])], axis=1).fillna(0).reset_index().set_index('flag')

    order_freq_summary_df = pd.DataFrame([
        {
            "production": 'ZZZQ1',
            "exch": exch,
            "order_time_max": p_df['order_cnt'].idxmax(),
            "cancel_time_max": p_df['cancel_cnt'].idxmax(),
            "order_freq": np.nanmax(p_df["order_cnt"]),
            "cancel_freq": np.nanmax(p_df["cancel_cnt"]),
            "order_and_cancel_freq": np.nanmax(p_df["order_cnt"] + p_df["cancel_cnt"]),

            "order_daily_sum": np.nansum(p_df["order_cnt"]),
            "cancel_daily_sum": np.nansum(p_df["cancel_cnt"]),
            "order_and_cancel_daily_sum": np.nansum(p_df["order_cnt"] + p_df["cancel_cnt"]),
        }
        for exch, p_df in order_freq_df.groupby("exch")
    ])

    order_freq_summary_df.to_csv(f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{curdate}/'
                                 f'OrderFreq/{curdate}_order_freq_1s_kafang.csv', index=False)


def freq_analysis_monthly(op_date):
    morning_timeindex_idx = pd.timedelta_range(start="09:30:00", end="11:30:00", freq="1s")
    after_timeindex_idx = pd.timedelta_range(start="13:00:00", end="15:00:00", freq="1s")
    timeindex_idx = morning_timeindex_idx.append(after_timeindex_idx)

    time_dict = {
        str(t)[7:15]: t for t in timeindex_idx
    }

    classifier = GeneralClassifier(time_dict)
    trade_cfg_df = read_trading_account_from_db(op_date)

    colo_list = get_colo_list(op_date)
    trader_list_cfg_df = get_trader_cfg(op_date)
    cfg_df = pd.merge(trader_list_cfg_df, trade_cfg_df[["colo",  "stid", "trade_acc", "production"]], on=["colo", "trade_acc"], how="inner")

    order_list = []
    for colo in colo_list:
        nas_parsed_dir = pta_task_profile.nas_parsed_dir()
        csv_path = os.path.join(nas_parsed_dir, colo, op_date, f"{colo}-worker-{op_date}.csv")
        if not os.path.exists(csv_path):
            nas_parsed_dir_old = pta_task_profile.nas_parsed_dir_old()
            csv_path = os.path.join(nas_parsed_dir_old, colo, op_date, f"{colo}-worker-{op_date}.csv")
        order_df = pd.read_csv(csv_path)
        if order_df.empty:
            continue
        order_df = order_df.sort_values('md_localtime', ascending=True)

        md_start_time = pd.Timedelta(order_df['md_exchtime'].iloc[0])
        md_end_time = pd.Timedelta(order_df['md_exchtime'].iloc[-1])
        md_local_start_time = order_df['md_localtime'].iloc[0]
        md_local_end_time = order_df['md_localtime'].iloc[-1]

        order_df['cancel_time_md'] = \
            (md_end_time - md_start_time).value / (md_local_end_time - md_local_start_time) * \
            (order_df['cancel_time'] - md_local_start_time) + md_start_time.value

        print(colo, colo_list.index(colo))
        order_df["colo"] = colo
        order_list.append(order_df)


    order_df = pd.concat(order_list, axis=0)
    order_df = order_df[order_df["send_err_time"].isnull()]
    #order_df = order_df[order_df["send_err_time"].isnull()]
    order_df = pd.merge(order_df, cfg_df[["colo", "stid", "production", "remote_colo", "exchange"]], on=["colo", "stid"], how="left")
    order_df["exch"] = order_df["insid"].apply(lambda t: t[:2])

    colo_filter_mask = order_df["remote_colo"].isnull() | \
                       (order_df["exchange"] == order_df["exch"])
    order_df = order_df[colo_filter_mask]

    order_df.rename({"orderid": "wf-orderid"}, inplace=True, axis=1)

    r_df = pd.read_excel(REPORT_PATH)
    r_df.rename({"内部代码": "production"}, axis=1, inplace=True)
    print(r_df.head(10))
    order_df = pd.merge(r_df, order_df, on="production", how="inner")

    order_df.reset_index(inplace=True, drop=True)
    order_df["md_exchtime"] = pd.to_timedelta(order_df["md_exchtime"])
    order_df["flag"] = classifier.classify(order_df["md_exchtime"])

    order_df["md_exchtime_shift_500"] = order_df["md_exchtime"] + pd.Timedelta(milliseconds=500)
    order_df["flag_shift"] = classifier.classify(order_df["md_exchtime_shift_500"])

    order_df_cancel = order_df[~order_df["cancel_time"].isnull()].copy(deep=True)
    order_df_cancel["cancel_time_md"] = pd.to_timedelta(order_df_cancel["cancel_time_md"])
    order_df_cancel["flag"] = classifier.classify(order_df_cancel["cancel_time_md"])

    order_df_cancel["cancel_time_md_shift_500"] = order_df_cancel["cancel_time_md"] + pd.Timedelta(milliseconds=500)
    order_df_cancel["flag_shift"] = classifier.classify(order_df_cancel["cancel_time_md_shift_500"])

    # bhprint(order_df)
    order_freq_df = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "order_cnt": len(flag_df),
            } for (p, flag, exch), flag_df in order_df.groupby(["原始产品名称", "flag", "exch"])
        ]).set_index(["production", "flag", "exch"])

    order_freq_df_cancel = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "cancel_cnt": len(flag_df),
            } for (p, flag, exch), flag_df in order_df_cancel.groupby(["原始产品名称", "flag", "exch"])
        ]).set_index(["production", "flag", "exch"])

    order_freq_df_shift = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "order_cnt_shift": len(flag_df),
            } for (p, flag, exch), flag_df in order_df.groupby(["原始产品名称", "flag_shift", "exch"])
        ]).set_index(["production", "flag", "exch"])

    order_freq_df_cancel_shift = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "exch": exch,
                "cancel_cnt_shift": len(flag_df),
            } for (p, flag, exch), flag_df in order_df_cancel.groupby(["原始产品名称", "flag_shift", "exch"])
        ]).set_index(["production", "flag", "exch"])

    order_freq_df = pd.concat([
        order_freq_df, order_freq_df_cancel,
        order_freq_df_shift, order_freq_df_cancel_shift], axis=1).fillna(0).reset_index()

    order_freq_summary_df = pd.DataFrame([
        {
            "production": p,

            "order_freq": np.nanmax(p_df["order_cnt"]),
            "cancel_freq": np.nanmax(p_df["cancel_cnt"]),
            "order_and_cancel_freq": np.nanmax(p_df["order_cnt"] + p_df["cancel_cnt"]),

            "order_freq_shift": np.nanmax(p_df["order_cnt_shift"]),
            "cancel_freq_shift": np.nanmax(p_df["cancel_cnt_shift"]),
            "order_and_cancel_freq_shift": np.nanmax(p_df["order_cnt_shift"] + p_df["cancel_cnt_shift"]),

            "order_daily_sum": np.nansum(p_df["order_cnt"]),
            "cancel_daily_sum": np.nansum(p_df["cancel_cnt"]),
            "order_and_cancel_daily_sum": np.nansum(p_df["order_cnt"] + p_df["cancel_cnt"]),
        }
        for p, p_df in order_freq_df.groupby("production")
    ]
    )
    return order_freq_summary_df


def order_freq_analysis_monthly(start_date, end_date, drop_date_list=None):
    start_trading_day = next_trading_day(pre_trading_day(start_date))
    trading_day = start_trading_day

    if drop_date_list is None:
        drop_date_list = []

    para_list = []
    while trading_day <= end_date:
        print(trading_day)
        if trading_day in drop_date_list:
            trading_day = next_trading_day(trading_day)
            continue

        para_list.append((trading_day,))
        trading_day = next_trading_day(trading_day)

    pl = Parallelism(processes=8)
    pl.run(freq_analysis_monthly, para_list)
    order_freq_df = pd.concat(pl.get_results(), axis=0)

    p_list = []
    for p, p_df in order_freq_df.groupby("production"):
        max_order_and_cancel_daily_sum = np.nanmax(p_df["order_and_cancel_daily_sum"])
        order_and_cancel_freq = np.nanmax(
            np.maximum(p_df["order_and_cancel_freq"], p_df["order_and_cancel_freq_shift"]))

        if order_and_cancel_freq < 100:
            range_freq = 'D-每秒申报笔数100笔以下'
        elif 100 <= order_and_cancel_freq < 300:
            range_freq = 'C-每秒申报笔数100笔至299笔'
        elif 300 <= order_and_cancel_freq < 500:
            range_freq = 'B-每秒申报笔数300笔至499笔'
        else:
            range_freq = 'A-每秒申报笔数500笔及以上'

        p_list.append(
            {
                "production": p,
                "max_order_and_cancel_daily_sum": max_order_and_cancel_daily_sum,
                "order_and_cancel_freq": order_and_cancel_freq,
                "range": range_freq
            }
        )

    p_df = pd.DataFrame(p_list)

    output_dir_w = f'{PLATFORM_PATH_DICT["y_path"]}TradingDailySummary/'
    if not os.path.exists(output_dir_w):
        os.makedirs(output_dir_w)

    output_dir = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{end_date}/OrderFreq/'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    p_df.to_excel(f"{output_dir}{start_date}-{end_date}_monthly_order_freq_summary.xlsx", index=False)
    p_df.to_excel(f"{output_dir_w}{start_date}-{end_date}_monthly_order_freq_summary.xlsx", index=False)


# def get_freq_order_daily(date, path_order):
#     df = pd.read_csv(path_order % (date, date))
#     df = df[df['production'].isin(WinterFallProductionList)]
#     df['撤单率'] = df['cancel_daily_sum'] / df['order_daily_sum']
#     df['Holdmv'] = df['production'].apply(lambda x: get_production_list_trading(date, x, 'holdmv'))
#     df['TradeValue'] = df['production'].apply(lambda x: calculate_pnl(date, x, ret_trade_value=True))
#     df['Turnover'] = df['TradeValue'] * df['Holdmv']
#     df['Date'] = date
#     return df
#
#
# def get_all_freq_order_infor(start_date, curdate):
#     path_order = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/%s/OrderFreq/%s_order_freq_1s.csv'
#     for date in get_trading_days(start_date, curdate):
#         if not os.path.exists(path_order % (date, date)):
#             print('不存在', date)
#
#     pool = mp.Pool(30)
#     conlist = pool.starmap(
#         get_freq_order_daily,
#         [(date, path_order) for date in get_trading_days(start_date, curdate)])
#     pool.close()
#     pool.join()
#
#     df = pd.concat(conlist, axis=0).sort_values(['Date', 'production']).reset_index(drop=True)
#     df = df.rename({'production': 'Product', 'order_freq': '秒最大发单数', 'cancel_freq': '秒最大撤单数',
#                     'order_and_cancel_freq': '秒最大发+撤单数', 'order_daily_sum': '日发单数',
#                     'cancel_daily_sum': '日撤单数', 'order_and_cancel_daily_sum': '日发+撤单数'}, axis='columns')
#     df = df.fillna('').style.set_properties(**{'text-align': 'center'}).set_table_styles(
#         [dict(selector='th', props=[('text-align', 'center')])]).set_table_styles([
#         {'selector': 'th', 'props': [('border', '1px solid black')]},
#         {'selector': 'td', 'props': [('border', '1px solid black')]}]).apply(
#         lambda x: ['background-color: {0}'.format('#FFB2DE') if np.abs(float(v) if v != '' else 0) > 300 else ''
#                    for v in x], axis=0, subset=['秒最大发单数', '秒最大撤单数', '秒最大发+撤单数']).apply(
#         lambda x: ['background-color: {0}'.format('#5B9B00') if np.abs(float(v) if v != '' else 0) >= 20000 else ''
#                    for v in x], axis=0, subset=['日发单数', '日撤单数', '日发+撤单数'])
#     df.to_excel(f'{LOG_TEMP_PATH}freq_order_info.xlsx', index=False)
#
#
# def plot_order_freq(start_date='20240317', end_date='20240416', production_list=None):
#     if production_list is None:
#         production_list = ['DC19',
#         'GDYPA',
#         'JQ11',
#         'JQ11B',
#         'JT1',
#         'ZJJT1',
#         'ZJJT2',
#         'ZXDC19',
#         'YP2',
#         'ZS9B'
#     ]
#     conlist = []
#     for date in get_trading_days(start_date, end_date):
#         df = pd.read_csv(
#             f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{date}/OrderFreq/{date}_order_freq_1s.csv')
#         df = df[df['production'].isin(production_list)]
#         df['Holdmv'] = df['production'].apply(lambda x: get_production_list_trading(date, x, ret_type='holdmv'))
#         df['Class'] = df['production'].apply(lambda x: production_2_index(x) + production_2_strategy(x))
#         df['Date'] = date
#         # bhprint(df)
#         conlist.append(df)
#     df = pd.concat(conlist, axis=0).drop(['order_time_max', 'cancel_time_max'], axis=1)
#     df = df[~df['Holdmv'].isna()]
#     df = df.rename({"order_freq": '最大发单数/s',
#                     'cancel_freq': '最大撤单数/s',
#                     'order_and_cancel_freq': '最大发+撤单数/s',
#                     'order_daily_sum': '日发单数',
#                     'cancel_daily_sum': '日撤单数',
#                     'order_and_cancel_daily_sum': '日发+撤单数'}, axis='columns').reset_index(drop=True)
#     df_all = df.copy(deep=True).sort_values('Holdmv', ascending=False)
#     plt.figure(figsize=(20, 16))
#     df_all_sz = df_all[df_all['exch'] == 'SZ']
#     df_all_sh = df_all[df_all['exch'] == 'SH']
#     sz_holdmv_list = np.log10(df_all_sz['Holdmv'].to_list())
#     sh_holdmv_list = np.log10(df_all_sh['Holdmv'].to_list())
#     size = 8
#
#     ax = plt.subplot(2, 2, 1)
#     plt.scatter(sz_holdmv_list, df_all_sz['最大发单数/s'].to_list(), s=size)
#     plt.scatter(sz_holdmv_list, df_all_sz['最大撤单数/s'].to_list(), s=size)
#     plt.scatter(sz_holdmv_list, df_all_sz['最大发+撤单数/s'].to_list(), s=size)
#     plt.legend(['最大发单数/s', '最大撤单数/s', '最大发+撤单数/s'])
#     ax.spines['top'].set_color('none')
#     ax.spines['right'].set_color('none')
#     # 移位置 设为原点相交
#     ax.yaxis.set_ticks_position('left')
#     ax.spines['left'].set_position(('data', 7.25))
#     ax.grid(True, linestyle='-.')
#     plt.title('SZ')
#
#     ax = plt.subplot(2, 2, 2)
#     plt.scatter(sh_holdmv_list, df_all_sh['最大发单数/s'].to_list(), s=size)
#     plt.scatter(sh_holdmv_list, df_all_sh['最大撤单数/s'].to_list(), s=size)
#     plt.scatter(sh_holdmv_list, df_all_sh['最大发+撤单数/s'].to_list(), s=size)
#     plt.legend(['最大发单数/s', '最大撤单数/s', '最大发+撤单数/s'])
#     ax.spines['top'].set_color('none')
#     ax.spines['right'].set_color('none')
#     # 移位置 设为原点相交
#     ax.yaxis.set_ticks_position('left')
#     ax.spines['left'].set_position(('data', 7.25))
#     ax.grid(True, linestyle='-.')
#     plt.title('SH')
#
#     ax = plt.subplot(2, 2, 3)
#     plt.scatter(sz_holdmv_list, df_all_sz['日发单数'].to_list(), s=size)
#     plt.scatter(sz_holdmv_list, df_all_sz['日撤单数'].to_list(), s=size)
#     plt.scatter(sz_holdmv_list, df_all_sz['日发+撤单数'].to_list(), s=size)
#     plt.legend(['日发单数', '日撤单数', '日发+撤单数'])
#     ax.spines['top'].set_color('none')
#     ax.spines['right'].set_color('none')
#     # 移位置 设为原点相交
#     ax.yaxis.set_ticks_position('left')
#     ax.spines['left'].set_position(('data', 7.25))
#     ax.grid(True, linestyle='-.')
#     plt.title('SZ')
#
#     ax = plt.subplot(2, 2, 4)
#     plt.scatter(sh_holdmv_list, df_all_sh['日发单数'].to_list(), s=size)
#     plt.scatter(sh_holdmv_list, df_all_sh['日撤单数'].to_list(), s=size)
#     plt.scatter(sh_holdmv_list, df_all_sh['日发+撤单数'].to_list(), s=size)
#     plt.legend(['日发单数', '日撤单数', '日发+撤单数'])
#     ax.spines['top'].set_color('none')
#     ax.spines['right'].set_color('none')
#     # 移位置 设为原点相交
#     ax.yaxis.set_ticks_position('left')
#     ax.spines['left'].set_position(('data', 7.25))
#     ax.grid(True, linestyle='-.')
#     plt.title('SH')
#     plt.suptitle(f'{start_date}-{end_date} | x=log10(holdmv):' + ','.join(production_list))
#     plt.tight_layout()
#     plt.savefig(f'{LOG_TEMP_PATH}{start_date}-{end_date}info_all.png')
#
#     df_all['Holdmv'] = np.round(df_all['Holdmv'] / 100000000, 2).astype('str') + ' kw'
#     df_all = df_all.astype('str').fillna('').style.set_properties(**{'text-align': 'center'}).set_table_styles(
#         [dict(selector='th', props=[('text-align', 'center')])]).set_table_styles([
#         {'selector': 'th', 'props': [('border', '1px solid black')]},
#         {'selector': 'td', 'props': [('border', '1px solid black')]}]).apply(
#         lambda x: ['background-color: {0}'.format('green') if np.abs(float(v) if v != '' else 0) > 300 else ''
#                    for v in x], axis=0, subset=['最大发单数/s', '最大撤单数/s', '最大发+撤单数/s']).apply(
#         lambda x: ['background-color: {0}'.format('green') if np.abs(float(v) if v != '' else 0) > 25000 else ''
#                    for v in x], axis=0, subset=['日发单数', '日撤单数', '日发+撤单数'])
#
#     df_all.to_excel(f'{LOG_TEMP_PATH}{start_date}-{end_date}info_all.xlsx', index=False, encoding='GBK')
#
#     df = df.groupby(['production', 'exch', 'Class']).mean().astype('int').reset_index()
#     df['Holdmv'] = np.round(df['Holdmv'] / 100000000, 2).astype('str') + ' kw'
#     df = df.astype('str').fillna('').style.set_properties(**{'text-align': 'center'}).set_table_styles(
#         [dict(selector='th', props=[('text-align', 'center')])]).set_table_styles([
#         {'selector': 'th', 'props': [('border', '1px solid black')]},
#         {'selector': 'td', 'props': [('border', '1px solid black')]}]).apply(
#         lambda x: ['background-color: {0}'.format('green') if np.abs(float(v) if v != '' else 0) > 300 else ''
#                    for v in x], axis=0, subset=['最大发单数/s', '最大撤单数/s', '最大发+撤单数/s']).apply(
#         lambda x: ['background-color: {0}'.format('green') if np.abs(float(v) if v != '' else 0) > 25000 else ''
#                    for v in x], axis=0, subset=['日发单数', '日撤单数', '日发+撤单数'])
#
#     df.to_excel(f'{LOG_TEMP_PATH}{start_date}-{end_date}_info.xlsx', index=False, encoding='GBK')


class GenerateProgrammedReportingInformationChange():
    def __init__(self, curdate):
        self.curdate = curdate
        self.path_template = f'{PLATFORM_PATH_DICT["y_path"]}8、信息披露进度/交易所股票程序化报备/券商有格式要求的信息报备表/'
        self.format_dict = {
            'gf': f'{self.path_template}广发证券股份有限公司/',
            'gj': f'{self.path_template}国泰君安证券股份有限公司/',
            'cms': f'{self.path_template}招商证券股份有限公司/',
            'cicc': f'{self.path_template}中国中金财富有限公司/',
            'csc': f'{self.path_template}中信建投证券股份有限公司/',
        }
        self.path_report = f'{PLATFORM_PATH_DICT["y_path"]}8、信息披露进度/交易所股票程序化报备/脚本/{curdate}/'
        self.email_cc_list = [
            'mofeiyue@centuryfrontier.com',
            'ritchguo@centuryfrontier.com',
        ]

    def format_report_date(self, date, date_type='-'):
        if not date_type:
            return date

        return datetime.datetime.strptime(date, '%Y%m%d').strftime(f'%Y{date_type}%m{date_type}%d')

    def generate_special_report(
            self, file_structure_dict, df_origin_report, broker_name, exchange, output_dir, date_type):
        new_infor_dict = {}
        for prod_name, trading_variety, order_freq, order_num in df_origin_report[
            ['账户名称', '交易品种', '账户最高申报速率（笔/秒）', '账户单日最高申报笔数']].values:
            new_infor_dict[prod_name.strip()] = {
                '交易品种': trading_variety,
                '申报速率': order_freq,
                '申报笔数': order_num,
                '券商名': broker_name,
            }

        broker_name_2_abbr = {
            '广发': 'gf',
            '招商': 'cms',
            '中信建投': 'csc',
        }

        start_row = file_structure_dict['start_row']
        workbook = openpyxl.load_workbook(
            self.format_dict[broker_name_2_abbr[broker_name]] + file_structure_dict['template_file'])
        sheet = workbook.active
        while True:
            product_name = str(
                sheet[f'{file_structure_dict["账户名称"]}{start_row}'].value).strip().replace(' ', '').replace('－', '-')
            if product_name in ['None', 'nan', '']:
                break
            if new_infor_dict.get(product_name, None) is not None:
                sheet[f'{file_structure_dict["报告日期"]}{start_row}'] = self.format_report_date(self.curdate, date_type)
                sheet[f'{file_structure_dict["交易品种"]}{start_row}'] = new_infor_dict[product_name]['交易品种']
                sheet[f'{file_structure_dict["账户最高申报速率"]}{start_row}'] = new_infor_dict[product_name][
                    '申报速率']
                sheet[f'{file_structure_dict["账户单日最高申报笔数"]}{start_row}'] = new_infor_dict[product_name][
                    '申报笔数']
                start_row += 1
                del new_infor_dict[product_name]
            else:
                sheet.delete_rows(start_row)

        if new_infor_dict:
            if exchange == 'sz':
                new_infor_dict['交易所'] = '深圳'
            else:
                new_infor_dict['交易所'] = '上海'
            print(f'{broker_name}_{exchange}_如下产品在模版中没有找到：', new_infor_dict)

        outputdir = output_dir + f'特殊格式报备表_{broker_name}/'
        if not os.path.exists(outputdir):
            os.makedirs(outputdir)

        workbook.save(outputdir + f'{broker_name}_特殊格式报备表_{exchange}_{self.curdate}.xlsx')
        return new_infor_dict

    def generate_special_report_gf(self, df_origin_report, output_dir, exchange='sz'):
        if exchange == 'sz':
            file_structure_dict = {
                '账户名称': 'D',
                '交易品种': 'Q',
                '报告日期': 'J',
                "账户最高申报速率": 'AA',
                "账户单日最高申报笔数": 'AB',
                'template_file': '世纪前沿-深圳20240401.xlsx',
                'start_row': 7,
            }
            date_type = '/'
        else:
            file_structure_dict = {
                '账户名称': 'C',
                '交易品种': 'R',
                '报告日期': 'I',
                "账户最高申报速率": 'AE',
                "账户单日最高申报笔数": 'AF',
                'template_file': '世纪前沿-上海20240401.xlsx',
                'start_row': 2,
            }
            date_type = ''
        res = self.generate_special_report(file_structure_dict, df_origin_report, '广发', exchange, output_dir, date_type)
        return res

    def generate_special_report_cms(self, df_origin_report, output_dir, exchange='sz'):
        if exchange == 'sz':
            file_structure_dict = {
                '账户名称': 'D',
                '交易品种': 'Q',
                '报告日期': 'J',
                "账户最高申报速率": 'AA',
                "账户单日最高申报笔数": 'AB',
                'template_file': '深圳证券交易所程序化交易投资者信息报告表_世纪前沿.xlsx',
                'start_row': 7,
            }
            date_type = ''
        else:
            file_structure_dict = {
                '账户名称': 'D',
                '交易品种': 'S',
                '报告日期': 'J',
                "账户最高申报速率": 'AF',
                "账户单日最高申报笔数": 'AG',
                'template_file': '上海证券交易所程序化交易投资者信息报告表_世纪前沿.xlsx',
                'start_row': 7
            }
            date_type = ''

        res = self.generate_special_report(file_structure_dict, df_origin_report, '招商', exchange, output_dir, date_type)
        return res

    def generate_special_report_csc(self, df_origin_report, output_dir, exchange='sz'):
        if exchange == 'sz':
            file_structure_dict = {
                '账户名称': 'D',
                '交易品种': 'Q',
                '报告日期': 'J',
                "账户最高申报速率": 'AA',
                "账户单日最高申报笔数": 'AB',
                'template_file': '深交所程序化交易信息报告表-世纪前沿-20240329.xlsx',
                'start_row': 7,
            }
            date_type = '/'
        else:
            file_structure_dict = {
                '账户名称': 'D',
                '交易品种': 'S',
                '报告日期': 'J',
                "账户最高申报速率": 'AF',
                "账户单日最高申报笔数": 'AG',
                'template_file': '上海证券交易所程序化交易投资者信息报告表-世纪前沿-20240329.xlsx',
                'start_row': 7,
            }
            date_type = '-'

        res = self.generate_special_report(file_structure_dict, df_origin_report, '中信建投', exchange, output_dir, date_type)
        return res

    def generate_special_report_cicc(self, df_origin_report, output_dir):
        file_structure_dict = {
            '账户名称': 'B',
            '报告日期': 'K',
            '沪市账户代码': 'E',
            '深市账户代码': 'F',
            '交易品种': 'T',
            "账户最高申报速率": 'AG',
            "账户单日最高申报笔数": 'AH',
            'template_file': '世纪前沿-程序化交易报备-更新上报终版（20240329）.xlsx',
            'start_row': 3,
        }

        broker_name = '中金财富'
        new_sz_infor_dict, new_sh_infor_dict = {}, {}
        for exchange, df_report_exch in df_origin_report.groupby('交易所'):
            for prod_name, trading_variety, order_freq, order_num in df_report_exch[
                ['账户名称', '交易品种', '账户最高申报速率（笔/秒）', '账户单日最高申报笔数']].values:
                if exchange == 'sz':
                    new_sz_infor_dict[prod_name.strip()] = {
                        '交易品种': trading_variety,
                        '申报速率': order_freq,
                        '申报笔数': order_num,
                        '券商名': broker_name,
                    }
                else:
                    new_sh_infor_dict[prod_name.strip()] = {
                        '交易品种': trading_variety,
                        '申报速率': order_freq,
                        '申报笔数': order_num,
                        '券商名': broker_name,
                    }

        start_row = file_structure_dict['start_row']
        workbook = openpyxl.load_workbook(self.format_dict['cicc'] + file_structure_dict['template_file'])
        sheet = workbook.active
        while True:
            product_name = str(
                sheet[f'{file_structure_dict["账户名称"]}{start_row}'].value).strip().replace(' ', '').replace('－', '-')
            sz_accid = str(sheet[f'{file_structure_dict["深市账户代码"]}{start_row}'].value).strip()
            if product_name in ['None', 'nan', '']:
                break

            if sz_accid not in ['None', 'nan', '']:
                if new_sz_infor_dict.get(product_name, None) is not None:
                    sheet[f'{file_structure_dict["报告日期"]}{start_row}'] = self.format_report_date(self.curdate, '/')
                    sheet[f'{file_structure_dict["交易品种"]}{start_row}'] = new_sz_infor_dict[product_name]['交易品种']
                    sheet[f'{file_structure_dict["账户最高申报速率"]}{start_row}'] = \
                        new_sz_infor_dict[product_name]['申报速率']
                    sheet[f'{file_structure_dict["账户单日最高申报笔数"]}{start_row}'] = \
                        new_sz_infor_dict[product_name]['申报笔数']
                    start_row += 1
                    del new_sz_infor_dict[product_name]
                else:
                    sheet.delete_rows(start_row)
            else:
                if new_sh_infor_dict.get(product_name, None) is not None:
                    sheet[f'{file_structure_dict["报告日期"]}{start_row}'] = self.format_report_date(self.curdate, '/')
                    sheet[f'{file_structure_dict["交易品种"]}{start_row}'] = new_sh_infor_dict[product_name]['交易品种']
                    sheet[f'{file_structure_dict["账户最高申报速率"]}{start_row}'] = \
                        new_sh_infor_dict[product_name]['申报速率']
                    sheet[f'{file_structure_dict["账户单日最高申报笔数"]}{start_row}'] = \
                        new_sh_infor_dict[product_name]['申报笔数']
                    start_row += 1
                    del new_sh_infor_dict[product_name]
                else:
                    sheet.delete_rows(start_row)

        if new_sz_infor_dict:
            new_sz_infor_dict['交易所'] = '深圳'
            print(f'{broker_name}_sz_如下产品在模版中没有找到：', new_sz_infor_dict)
        if new_sh_infor_dict:
            new_sh_infor_dict['交易所'] = '上海'
            print(f'{broker_name}_sh_如下产品在模版中没有找到：', new_sh_infor_dict)

        outputdir = output_dir + f'特殊格式报备表_{broker_name}/'
        if not os.path.exists(outputdir):
            os.makedirs(outputdir)

        workbook.save(outputdir + f'{broker_name}_特殊格式报备表_沪深_{self.curdate}.xlsx')
        return new_sz_infor_dict, new_sh_infor_dict

    def generate_special_report_gj(self, df_origin_report, output_dir):
        file_structure_dict = {
            '账户名称': 'I',
            '市场': 'F',
            '报告日期': 'P',
            '交易品种': 'Y',
            "账户最高申报速率": 'AL',
            "账户单日最高申报笔数": 'AM',
            'template_file': '海南世纪前沿私募基金管理有限公司-程序化交易投资者信息报告表(变更20240305).xlsx',
            'start_row': 5,
        }

        broker_name = '国泰君安'
        new_sz_infor_dict, new_sh_infor_dict = {}, {}
        for exchange, df_report_exch in df_origin_report.groupby('交易所'):
            for prod_name, trading_variety, order_freq, order_num in df_report_exch[
                ['账户名称', '交易品种', '账户最高申报速率（笔/秒）', '账户单日最高申报笔数']].values:
                if exchange == 'sz':
                    new_sz_infor_dict[prod_name.strip()] = {
                        '交易品种': trading_variety,
                        '申报速率': order_freq,
                        '申报笔数': order_num,
                        '券商名': broker_name,
                    }
                else:
                    new_sh_infor_dict[prod_name.strip()] = {
                        '交易品种': trading_variety,
                        '申报速率': order_freq,
                        '申报笔数': order_num,
                        '券商名': broker_name,
                    }

        start_row = file_structure_dict['start_row']
        workbook = openpyxl.load_workbook(self.format_dict['gj'] + file_structure_dict['template_file'])
        sheet = workbook.active
        while True:
            product_name = str(
                sheet[f'{file_structure_dict["账户名称"]}{start_row}'].value).strip().replace(' ', '').replace('－', '-')
            sz_accid = str(sheet[f'{file_structure_dict["市场"]}{start_row}'].value).strip()
            if product_name in ['None', 'nan', '']:
                break

            if sz_accid == '深圳':
                if new_sz_infor_dict.get(product_name, None) is not None:
                    sheet[f'{file_structure_dict["报告日期"]}{start_row}'] = self.format_report_date(self.curdate, '-')
                    sheet[f'{file_structure_dict["交易品种"]}{start_row}'] = new_sz_infor_dict[product_name]['交易品种']
                    sheet[f'{file_structure_dict["账户最高申报速率"]}{start_row}'] = \
                        new_sz_infor_dict[product_name]['申报速率']
                    sheet[f'{file_structure_dict["账户单日最高申报笔数"]}{start_row}'] = \
                        new_sz_infor_dict[product_name]['申报笔数']
                    start_row += 1
                    del new_sz_infor_dict[product_name]
                else:
                    sheet.delete_rows(start_row)
            else:
                if new_sh_infor_dict.get(product_name, None) is not None:
                    sheet[f'{file_structure_dict["报告日期"]}{start_row}'] = self.format_report_date(self.curdate, '-')
                    sheet[f'{file_structure_dict["交易品种"]}{start_row}'] = new_sh_infor_dict[product_name]['交易品种']
                    sheet[f'{file_structure_dict["账户最高申报速率"]}{start_row}'] = \
                        new_sh_infor_dict[product_name]['申报速率']
                    sheet[f'{file_structure_dict["账户单日最高申报笔数"]}{start_row}'] = \
                        new_sh_infor_dict[product_name]['申报笔数']
                    start_row += 1
                    del new_sh_infor_dict[product_name]
                else:
                    sheet.delete_rows(start_row)

        if new_sz_infor_dict:
            new_sz_infor_dict['交易所'] = '深圳'
            print(f'{broker_name}_sz_如下产品在模版中没有找到：', new_sz_infor_dict)
        if new_sh_infor_dict:
            new_sh_infor_dict['交易所'] = '上海'
            print(f'{broker_name}_sh_如下产品在模版中没有找到：', new_sh_infor_dict)

        outputdir = output_dir + f'特殊格式报备表_{broker_name}/'
        if not os.path.exists(outputdir):
            os.makedirs(outputdir)

        workbook.save(outputdir + f'{broker_name}_特殊格式报备表_沪深_{self.curdate}.xlsx')
        return new_sz_infor_dict, new_sh_infor_dict

    def generate_special_report_from_origin(self, df_report_sz, df_report_sh, output_dir):
        df_report_sz = df_report_sz.copy(deep=True)
        df_report_sh = df_report_sh.copy(deep=True)
        df_report_sz['交易所'] = 'sz'
        df_report_sh['交易所'] = 'sh'

        broker_infor_dict = {
            'cms': '招商证券股份有限公司',
            'csc': '中信建投证券股份有限公司',
            'gf': '广发证券股份有限公司',
        }
        res_infor_dict_list = []
        for broker in broker_infor_dict:
            func_generate_special_report = eval(f'self.generate_special_report_{broker}')

            df_report_sz_broker = df_report_sz[df_report_sz['会员/其他机构名称'] == broker_infor_dict[broker]].copy(deep=True)
            if not df_report_sz_broker.empty:
                sz_res = func_generate_special_report(df_report_sz_broker, output_dir, 'sz')
                if sz_res:
                    res_infor_dict_list.append(sz_res)

            df_report_sh_broker = df_report_sh[df_report_sh['会员/其他机构名称'] == broker_infor_dict[broker]].copy(deep=True)
            if not df_report_sh_broker.empty:
                sh_res = func_generate_special_report(df_report_sh_broker, output_dir, 'sh')
                if sh_res:
                    res_infor_dict_list.append(sh_res)

        broker_infor_dict = {
            'gj': '国泰君安证券股份有限公司',
            'cicc': '中国中金财富有限公司',
        }
        for broker in broker_infor_dict:
            func_generate_special_report = eval(f'self.generate_special_report_{broker}')

            df_report_broker = pd.concat([
                df_report_sz[df_report_sz['会员/其他机构名称'] == broker_infor_dict[broker]].copy(deep=True),
                df_report_sh[df_report_sh['会员/其他机构名称'] == broker_infor_dict[broker]].copy(deep=True)
            ], axis=0)
            if not df_report_broker.empty:
                sz_res, sh_res = func_generate_special_report(df_report_broker, output_dir)
                if sz_res:
                    res_infor_dict_list.append(sz_res)
                if sh_res:
                    res_infor_dict_list.append(sh_res)

        if res_infor_dict_list:
            conlist = []
            for res in res_infor_dict_list:
                exchange = res['交易所']
                del res['交易所']
                df_res = pd.DataFrame.from_dict(res, orient='index')
                df_res['交易所'] = exchange

                conlist.append(df_res)
            df_res = pd.concat(conlist, axis=0).reset_index().rename({'index': '产品名称'}, axis='columns')
            print(df_res)
            df_res.to_excel(
                f'{self.path_report}{self.curdate}_特殊格式报备表_产品名称_缺失情况汇总.xlsx', index=False)

    def generate_programmed_reporting_information_change_monthly_report(self, curdate):
        path_infor = f'{PLATFORM_PATH_DICT["y_path"]}8、信息披露进度/交易所股票程序化报备/每月更新_券商接收邮箱.xlsx'
        df_email_infor = pd.read_excel(path_infor)
        df_email_infor['内部代码'] = df_email_infor['内部代码'].fillna('所有产品').str.strip()
        df_email_infor = df_email_infor.sort_values('内部代码')

        df_report_sz = pd.read_excel(self.path_report + f'{curdate}_compare_file_sz.xlsx', dtype='str')
        df_report_sh = pd.read_excel(self.path_report + f'{curdate}_compare_file_sh.xlsx', dtype='str')

        self.generate_special_report_from_origin(df_report_sz, df_report_sh, self.path_report)

        delete_col_list = list({'Proudct', 'Unnamed: 0', '开户营业部', 'compare_col', '交易状态'}.intersection(
            set(df_report_sz.columns.to_list())))

        broker_2_email_info, infor_list = {}, []
        for broker, df_brkr in df_email_infor.groupby('经纪商'):
            df_sz_broker = df_report_sz[df_report_sz['会员/其他机构名称'] == broker.strip()]
            df_sh_broker = df_report_sh[df_report_sh['会员/其他机构名称'] == broker.strip()]
            product_list = list(df_sz_broker['Proudct'].to_list() + df_sh_broker['Proudct'].to_list())

            email_2_prod_dict = {}
            for email, prod_l in df_brkr[['邮箱', '内部代码']].values:
                email = email.replace('；', ';').replace(' ', '')
                if '所有产品' in prod_l:
                    email_prod_list = product_list
                else:
                    email_prod_list = prod_l.replace('；', ';').replace(' ', '').split(';')
                    product_list = list(set(product_list) - set(email_prod_list))

                if email_2_prod_dict.get(email, None) is None:
                    email_2_prod_dict[email] = email_prod_list
                else:
                    email_prod_list = deepcopy(email_prod_list + email_2_prod_dict[email])
                    email_2_prod_dict[email] = email_prod_list

            broker_2_email_info[broker] = email_2_prod_dict
            for email in email_2_prod_dict.keys():
                infor_list.append([
                    broker,
                    email,
                    ';'.join(sorted(self.email_cc_list)),
                    ';'.join(sorted(email_2_prod_dict[email])),
                    f'{broker}_Exch_{email}',
                ])

                df_prod_sz = df_sz_broker[df_sz_broker['Proudct'].isin(email_2_prod_dict[email])]
                if not df_prod_sz.empty:
                    df_prod_sz = df_prod_sz.drop(delete_col_list, axis=1)
                    file_name_send = f'{broker}_深圳_{email}'
                    df_prod_sz.to_excel(f'{self.path_report}{file_name_send}.xlsx', index=False)

                df_prod_sh = df_sh_broker[df_sh_broker['Proudct'].isin(email_2_prod_dict[email])]
                if not df_prod_sh.empty:
                    df_prod_sh = df_prod_sh.drop(delete_col_list, axis=1)
                    file_name_send = f'{broker}_上海_{email}'
                    df_prod_sh.to_excel(f'{self.path_report}{file_name_send}.xlsx', index=False)

        df_check_result = pd.DataFrame(infor_list, columns=['经纪商', '邮箱', '邮箱cc', '产品名', '文件名'])
        df_check_result.to_csv(f'{self.path_report}check_total_information.csv', index=False, encoding='GBK')

        cfg_path_report = f'{self.path_report}{curdate}_broker_email_report_infor.cfg'
        with open(cfg_path_report, 'w', encoding='GBK') as f:
            f.write(libconf.dumps(broker_2_email_info))
        print(broker_2_email_info)

    def send_email_programmed_reporting_information_change_monthly_report(self, curdate):
        cfg_path_report = f'{self.path_report}{curdate}_broker_email_report_infor.cfg'
        broker_email_info_dict = {}
        with open(cfg_path_report, 'r', encoding='GBK') as fr:
            for line in fr.read().replace('\n', '').split(';};'):
                if not line.strip():
                    continue
                broker, email_str = line.split('={')
                email_info_dict = {}

                for email_info in email_str.split('; '):
                    print(broker, email_info)
                    if not email_info.strip():
                        continue
                    email, prod_list = email_info.split('=')
                    email_info_dict[email.strip()] = eval(prod_list.strip())
                broker_email_info_dict[broker.strip()] = email_info_dict

        for broker_name in broker_email_info_dict:
            for email in broker_email_info_dict[broker_name]:
                email_list = email.split(';')
                attchment_dir = []
                sz_path_file = f'{self.path_report}{broker_name}_深圳_{email}.xlsx'
                sh_path_file = f'{self.path_report}{broker_name}_上海_{email}.xlsx'
                if os.path.exists(sz_path_file):
                    attchment_dir.append(sz_path_file)

                if os.path.exists(sh_path_file):
                    attchment_dir.append(sh_path_file)

                if not attchment_dir:
                    print('不存在\n', sz_path_file, sh_path_file)
                    continue

                print(sz_path_file, sh_path_file, email_list)

                ase = AutoSendEmailTxt(
                    curdate=curdate,
                    content_txt="您好！\n\t附件为海南世纪前沿程序化交易投资者信息报告表，上个月账户信息有变更，请及时帮上报沪深交易所，麻烦上报交易所后，请回复邮件“已上报交易所”。\n\t谢谢！",
                    subject=f'海南世纪前沿_{curdate}_程序化交易投资者信息报告表（变更）报备_{broker_name}',
                    receivers=email_list,
                    receivers_cc=self.email_cc_list,
                    attchmentdir=attchment_dir,
                    sender='data@centuryfrontier.com',
                    sender_pwd='Cfi-0001',
                )
                ase.send_email()


def statistics_cancel_ratio_second(dict_info=None):
    if dict_info is None:
        dict_info = {}
        for fpath in list(
                Path(f'{PLATFORM_PATH_DICT["z_path"]}Trading/POB_file/pob_files/').glob('*-POB-2024*.csv')):
            production, _, date = fpath.stem.split('.')[0].split('-')
            dict_info[production] = date

    infor_list = []
    pta_task_profile = PtaTaskProfile()
    for production in dict_info:
        date = dict_info[production]

        colo_name = production_2_colo(production)
        path_colo = f'{PLATFORM_PATH_DICT["z_path"]}ProcessedData/PTA/parsed_log/{colo_name}/{date}/'
        print(production, path_colo)
        if not os.path.exists(path_colo):
            continue
        alpha_dict = get_trading_accounts_paras_dict('productions', path_colo)
        if alpha_dict.get(production, None) is None:
            continue

        alpha_i = alpha_dict[production]

        df_worker = pd.read_csv(f'{path_colo}{colo_name}-worker-{date}.csv')
        df_worker = df_worker[df_worker['stid'].astype('int') == alpha_i]
        machine_cpu_freq = pta_task_profile.machine_cpu_freq(colo_name, date)
        df_worker['colo'] = colo_name
        df_worker['cancel_time'] /= machine_cpu_freq
        df_worker['trader_sent'] /= machine_cpu_freq

        if production in DUALCENTER_PRODUCTION:
            colo_name = colo_name.replace('sz', 'sh')
            path_colo = f'{PLATFORM_PATH_DICT["z_path"]}ProcessedData/PTA/parsed_log/{colo_name}/{date}/'
            df_worker_sh = pd.read_csv(f'{path_colo}{colo_name}-worker-{date}.csv')
            df_worker_sh = df_worker_sh[df_worker_sh['stid'] == alpha_i]
            df_worker_sh = df_worker_sh[df_worker_sh['insid'].apply(lambda x: x[2] == '6')]
            df_worker_sh['colo'] = colo_name

            machine_cpu_freq = pta_task_profile.machine_cpu_freq(colo_name, date)
            df_worker_sh['cancel_time'] /= machine_cpu_freq
            df_worker_sh['trader_sent'] /= machine_cpu_freq

            df_worker = df_worker[df_worker['insid'].apply(lambda x: x[2] != '6')]
            df_worker = pd.concat([df_worker, df_worker_sh], axis=0)

        df_worker = df_worker[df_worker['canceled']]
        df_worker.to_csv(f'{LOG_TEMP_PATH}worker.csv')
        df_worker = df_worker[['insid', 'stid', 'direction', 'md_exchtime', 'canceled', 'trader_sent', 'cancel_time', 'colo']]
        df_worker['diff'] = (df_worker['cancel_time'] - df_worker['trader_sent']) / 1000
        df_worker.to_csv(f'{LOG_TEMP_PATH}{date}_{production}_worker.csv')
        infor_list.append([
            production, date, len(df_worker[df_worker['diff'] < 1])
        ])

    df_infor = pd.DataFrame(infor_list, columns=['Product', 'Date', 'Num']).sort_values('Num', ascending=False)
    df_infor = df_infor.reset_index(drop=True)
    df_infor.to_csv(f'{LOG_TEMP_PATH}more_than_1s.csv', index=False)
    print(df_infor)


def get_order_infor_data_by_product(op_date, product):
    morning_timeindex_idx = pd.timedelta_range(start="09:30:00", end="11:30:00", freq="1s")
    after_timeindex_idx = pd.timedelta_range(start="13:00:00", end="15:00:00", freq="1s")
    timeindex_idx = morning_timeindex_idx.append(after_timeindex_idx)

    time_dict = {
        str(t)[7:15]: t for t in timeindex_idx
    }

    classifier = GeneralClassifier(time_dict)
    trade_cfg_df = read_trading_account_from_db(op_date)
    trader_list_cfg_df = get_trader_cfg(op_date)

    cfg_df = pd.merge(trader_list_cfg_df, trade_cfg_df[["colo", "stid", "trade_acc", "production"]],
                      on=["colo", "trade_acc"], how="inner")
    cfg_df = cfg_df[cfg_df['production'] == product]

    pta_task_profile = PtaTaskProfile()
    colo_2_machine_cpu_freq = {
        colo: pta_task_profile.machine_cpu_freq(colo, op_date) for colo in cfg_df['colo'].unique()}
    cfg_df['cpu_freq'] = cfg_df['colo'].apply(lambda x: colo_2_machine_cpu_freq[x])

    colo = production_2_colo(product)
    colo_list = [colo, production_2_colo_sh(product, colo_sz=colo)]
    order_list = []
    for colo in colo_list:
        nas_parsed_dir = pta_task_profile.nas_parsed_dir()
        csv_path = os.path.join(nas_parsed_dir, colo, op_date, f"{colo}-worker-{op_date}.csv")
        if not os.path.exists(csv_path):
            nas_parsed_dir_old = pta_task_profile.nas_parsed_dir_old()
            csv_path = os.path.join(nas_parsed_dir_old, colo, op_date, f"{colo}-worker-{op_date}.csv")
        order_df = pd.read_csv(csv_path)
        print(colo, op_date, colo_list.index(colo))
        order_df["colo"] = colo
        order_list.append(order_df)

    order_df = pd.concat(order_list, axis=0)
    order_df = order_df[order_df["send_err_time"].isnull()]
    order_df = pd.merge(order_df, cfg_df[
        ["colo", "stid", "production", "remote_colo", "exchange", "cpu_freq"]], on=["colo", "stid"], how="left")
    order_df["exch"] = order_df["insid"].apply(lambda t: t[:2])
    colo_filter_mask = order_df["remote_colo"].isnull() | (order_df["exchange"] == order_df["exch"])
    order_df = order_df[colo_filter_mask]
    order_df = order_df[order_df['production'] == product]

    order_df.rename({"orderid": "wf-orderid"}, inplace=True, axis=1)
    order_df['cancel_wait_time'] = (order_df['cancel_time'] - order_df['trader_sent']) / order_df['cpu_freq'] / 1000
    order_df['cancel_in_1s'] = (order_df['cancel_wait_time'] < 1).astype('int')
    order_df.reset_index(inplace=True, drop=True)
    order_df["md_exchtime"] = pd.to_timedelta(order_df["md_exchtime"])
    order_df["flag"] = classifier.classify(order_df["md_exchtime"])
    order_df["flag"] = order_df["flag"].apply(lambda x: float(str(x).replace(':', '')))
    order_df.to_csv(f'{LOG_TEMP_PATH}{op_date}_{product}_order.csv', index=False)
    hprint(order_df)


def analysis_intraday_freq(curdate, product, output_dir=None):
    if output_dir is None:
        output_dir = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{curdate}/OrderFreq/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if isinstance(product, str):
        product = [product]

    morning_timeindex_idx = pd.timedelta_range(start="09:30:00", end="11:30:00", freq="1s")
    after_timeindex_idx = pd.timedelta_range(start="13:00:00", end="15:00:00", freq="1s")
    timeindex_idx = morning_timeindex_idx.append(after_timeindex_idx)
    time_dict = {
        str(t)[7:15]: t for t in timeindex_idx
    }
    classifier = GeneralClassifier(time_dict)

    morning_timeindex_idx_min = pd.timedelta_range(start="09:30:00", end="11:30:00", freq="1min")
    after_timeindex_idx_min = pd.timedelta_range(start="13:00:00", end="15:00:00", freq="1min")
    timeindex_idx_min = morning_timeindex_idx_min.append(after_timeindex_idx_min)
    time_dict_min = {
        str(t)[7:15]: t for t in timeindex_idx_min
    }
    classifier_min = GeneralClassifier(time_dict_min)

    trade_cfg_df = read_trading_account_from_db(curdate)

    colo_list = []
    for prod in product:
        colo = production_2_colo(prod)
        colo_list.append(colo)
        if prod in DUALCENTER_PRODUCTION:
            colo_list.append(production_2_colo_sh(prod, colo))

    trader_list_cfg_df = get_trader_cfg(curdate)
    cfg_df = pd.merge(trader_list_cfg_df, trade_cfg_df[["colo", "stid", "trade_acc", "production"]],
                      on=["colo", "trade_acc"], how="inner")
    cfg_df = cfg_df[cfg_df['colo'].isin(colo_list)]

    pta_task_profile = PtaTaskProfile()
    colo_2_machine_cpu_freq = {
        colo: pta_task_profile.machine_cpu_freq(colo, curdate) for colo in cfg_df['colo'].unique()}
    cfg_df['cpu_freq'] = cfg_df['colo'].apply(lambda x: colo_2_machine_cpu_freq[x])

    order_list = []
    for colo in colo_list:
        nas_parsed_dir = pta_task_profile.nas_parsed_dir()
        csv_path = os.path.join(nas_parsed_dir, colo, curdate, f"{colo}-worker-{curdate}.csv")
        if not os.path.exists(csv_path):
            nas_parsed_dir_old = pta_task_profile.nas_parsed_dir_old()
            csv_path = os.path.join(nas_parsed_dir_old, colo, curdate, f"{colo}-worker-{curdate}.csv")
        order_df = pd.read_csv(csv_path)
        if order_df.empty:
            continue
        order_df = order_df.sort_values('md_localtime', ascending=True)

        md_start_time = pd.Timedelta(order_df['md_exchtime'].iloc[0])
        md_end_time = pd.Timedelta(order_df['md_exchtime'].iloc[-1])
        md_local_start_time = order_df['md_localtime'].iloc[0]
        md_local_end_time = order_df['md_localtime'].iloc[-1]

        order_df['cancel_time_md'] = \
            (md_end_time - md_start_time).value / (md_local_end_time - md_local_start_time) * \
            (order_df['cancel_time'] - md_local_start_time) + md_start_time.value

        print(colo, curdate, colo_list.index(colo))
        order_df["colo"] = colo
        order_list.append(order_df)

    order_df = pd.concat(order_list, axis=0)
    order_df = order_df[order_df["send_err_time"].isnull()]
    order_df = pd.merge(order_df, cfg_df[
        ["colo", "stid", "production", "remote_colo", "exchange", "cpu_freq"]], on=["colo", "stid"], how="left")
    order_df["exch"] = order_df["insid"].apply(lambda t: t[:2])
    order_df = order_df[order_df['production'].isin(product)]
    colo_filter_mask = order_df["remote_colo"].isnull() | (order_df["exchange"] == order_df["exch"])
    order_df = order_df[colo_filter_mask]

    order_df.rename({"orderid": "wf-orderid"}, inplace=True, axis=1)
    order_df['cancel_wait_time'] = (order_df['cancel_time'] - order_df['trader_sent']) / order_df['cpu_freq'] / 1000
    order_df['cancel_in_1s'] = (order_df['cancel_wait_time'] < 1).astype('int')

    order_df.reset_index(inplace=True, drop=True)
    order_df["md_exchtime"] = pd.to_timedelta(order_df["md_exchtime"])
    order_df["flag"] = classifier.classify(order_df["md_exchtime"])
    order_df["flag_min"] = classifier_min.classify(order_df["md_exchtime"])

    order_df_cancel = order_df[~order_df["cancel_time"].isnull()].copy(deep=True)
    order_df_cancel["cancel_time_md"] = pd.to_timedelta(order_df_cancel["cancel_time_md"])
    order_df_cancel["flag"] = classifier.classify(order_df_cancel["cancel_time_md"])
    order_df_cancel["flag_min"] = classifier_min.classify(order_df_cancel["cancel_time_md"])

    order_freq_df = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "flag_min": flag_min,
                "exch": exch,
                "code": code,
                "direction": direction,
                "order_cnt": len(flag_df),
                "volume": flag_df['trade_volume'].sum(),
                "value": flag_df['trade_turnover'].sum(),
            } for (p, flag, flag_min, exch, direction, code), flag_df in order_df.groupby(
            ["production", "flag", "flag_min", "exch", "direction", "insid"])
        ]).set_index(["production", "flag", "flag_min", "exch", "direction", "code"])

    order_freq_df_cancel = pd.DataFrame(
        [
            {
                "production": p,
                "flag": flag,
                "flag_min": flag_min,
                "exch": exch,
                "code": code,
                "direction": direction,
                "cancel_cnt": len(flag_df),
                "cancel_volume": (flag_df['volume'] - flag_df['trade_volume']).sum(),
                "cancel_value": (flag_df['volume'] * flag_df['price'] - flag_df['trade_turnover']).sum(),
                "cancel_num_in_1s": flag_df["cancel_in_1s"].sum(),
            } for (p, flag, flag_min, exch, direction, code), flag_df in order_df_cancel.groupby(
            ["production", "flag", "flag_min", "exch", "direction", "insid"])
        ]).set_index(["production", "flag", "flag_min", "exch", "direction", "code"])

    order_freq_df = pd.concat([
        order_freq_df, order_freq_df_cancel], axis=1).fillna(0).reset_index()
    order_freq_df['total_cnt'] = order_freq_df["order_cnt"] + order_freq_df["cancel_cnt"]

    order_freq_df_1s = order_freq_df.groupby(["production", "flag", "flag_min", "exch"])[
        ['order_cnt', 'cancel_cnt', 'total_cnt']].sum().reset_index()
    order_freq_df_1s['Date'] = curdate

    order_freq_df_1s_all = order_freq_df_1s.groupby(["flag", "flag_min", "exch"])[
        ['order_cnt', 'cancel_cnt', 'total_cnt']].sum().reset_index()
    order_freq_df_1s_all['Date'] = curdate

    order_freq_df_1s.to_csv(f'{LOG_TEMP_PATH}{curdate}_1s_{"-".join(product)}.csv', index=False)
    order_freq_df_1s_all.to_csv(f'{LOG_TEMP_PATH}{curdate}_1s_{"-".join(product)}.csv', index=False)

    # order_freq_df_copy = order_freq_df.copy(deep=True)
    # order_freq_df_copy['code'] = order_freq_df_copy['code'].apply(lambda x: x[2:])
    # order_freq_df_copy['flag'] = order_freq_df_copy['flag'].apply(lambda x: int(x.replace(':', '')))
    # df_1min = get_n_min_stock_daily_data(curdate, '1min')[['time', 'code', 'volume', 'turnover', 'high', 'low', 'close']].rename(
    #     {'volume': 'market_vol'}, axis='columns')
    # df_1min['change'] = (df_1min['high'] - df_1min['low']) / df_1min['close']
    # order_freq_df_copy = pd.merge(order_freq_df_copy, df_1min, left_on=['flag', 'code'], right_on=['time', 'code'], how='left')
    # order_freq_df_copy['ratio'] = order_freq_df_copy['volume'] / order_freq_df_copy['market_vol']

    index_time = [str(ti)[7:15] for ti in timeindex_idx]
    index_time_min = [str(ti)[7:15] for ti in timeindex_idx_min]
    print(len(index_time))

    order_freq_df.columns.name = None
    order_freq_df.index.name = None
    for prod in product:
        plt.figure(figsize=(20, 12))
        for subi, ((exch, direction), df_sub) in enumerate(order_freq_df.groupby(['exch', 'direction'])):
            direction = 'Long' if direction == 0 else 'Short'
            hprint(df_sub)
            df_sub_min = df_sub.groupby('flag_min')[['order_cnt', 'cancel_cnt', 'total_cnt']].sum().fillna(0).rename(
                {'order_cnt': '发单数', 'cancel_cnt': '撤单数', 'total_cnt': '总发单数'}, axis='columns')
            hprint(df_sub_min)

            df_sub_min = df_sub_min.reindex(index_time_min, fill_value=0)
            ax = plt.subplot(4, 3, subi * 2 + 1)
            df_sub_min = rename_df_columns_name(df_sub_min, mode='sum/max')
            df_sub_min.plot.bar(ax=ax)
            format_ax(ax)
            ax.xaxis.set_major_locator(MultipleLocator(15))
            plt.title(f'{exch}-{direction}')

            ax = plt.subplot(4, 3, subi * 2 + 2)
            df_sub = df_sub.groupby('flag')[['order_cnt', 'cancel_cnt', 'total_cnt']].sum().fillna(0).rename(
                {'order_cnt': '发单数', 'cancel_cnt': '撤单数', 'total_cnt': '总发单数'}, axis='columns')
            df_sub = df_sub.reindex(index_time, fill_value=0)
            df_sub = np.cumsum(df_sub, axis=0)
            df_sub.plot(ax=ax)
            ax.xaxis.set_major_locator(MultipleLocator(900))
            format_ax(ax)
            plt.title(f'{exch}-{direction}')
            # ax.xaxis.set_major_locator(MultipleLocator(15))
            #
            # xticks_list, _ = plt.xticks()
            # print(xticks_list)
            # plt.xticks(xticks_list, [''] + [index_time[int(xt)] for xt in xticks_list[1:-1] + ['']], rotation=30)

        df_sub_min = order_freq_df.groupby('flag_min')[['order_cnt', 'cancel_cnt', 'total_cnt']].sum().fillna(0).rename(
            {'order_cnt': '发单数', 'cancel_cnt': '撤单数', 'total_cnt': '总发单数'}, axis='columns')
        df_sub_min = df_sub_min.reindex(index_time_min, fill_value=0)
        ax = plt.subplot(4, 3, 9)
        df_sub_min = rename_df_columns_name(df_sub_min, mode='sum/max')
        df_sub_min.plot.bar(ax=ax)
        ax.xaxis.set_major_locator(MultipleLocator(15))
        format_ax(ax)
        plt.title(f'total')

        ax = plt.subplot(4, 3, 10)
        df_sub = order_freq_df.groupby('flag')[['order_cnt', 'cancel_cnt', 'total_cnt']].sum().fillna(0).rename(
            {'order_cnt': '发单数', 'cancel_cnt': '撤单数', 'total_cnt': '总发单数'}, axis='columns')
        df_sub = df_sub.reindex(index_time, fill_value=0)
        df_sub = np.cumsum(df_sub, axis=0)
        df_sub.plot(ax=ax)
        format_ax(ax)
        ax.xaxis.set_major_locator(MultipleLocator(900))
        plt.title(f'total')

        ax = plt.subplot(4, 3, 11)
        plot_temp = order_freq_df[(order_freq_df['exch'] == 'SH') & (order_freq_df['direction'] == 1)]['total_cnt'].to_list()
        plt.hist(plot_temp)
        plt.legend([f'Max: {np.max(plot_temp)}'])
        format_ax(ax)

        # ax.xaxis.set_major_locator(MultipleLocator(15))
        plt.grid(True, linestyle='--')

        plt.suptitle(f'{curdate} | {prod}, delta=15min')
        plt.tight_layout()
        order_freq_df.to_csv(f'{output_dir}{curdate}_{prod}_order_infor.csv', index=False)
        plt.savefig(f'{output_dir}{curdate}_{prod}_order_infor.png')
    # plt.show()


def freq_analysis_t0_status(date, output_dir=None, num_pool=30):
    if output_dir is None:
        output_dir = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{date}/OrderFreq/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # colo_list = ['cicc-sz-4', 'cicc-sh-4']
    code_list = get_code_list(date, 'SH-')
    print(code_list)

    trade_info_list = pool_starmap_multiprocessing(get_trade_info, [(code, date) for code in code_list], num_pool)
    order_list = []
    for df_trade_info in trade_info_list:
        if df_trade_info.empty:
            continue
        df_trade_info = trade_agg_func(df_trade_info)
        order_list.append(df_trade_info)

    order_df = pd.concat(order_list, axis=0)
    order_df.reset_index(inplace=True, drop=True)
    order_df["md_exchtime"] = pd.to_timedelta(order_df["md_exchtime"])
    order_df["exch"] = order_df["insid"].apply(lambda t: t[:2])
    order_df['canceled'] = order_df['volume'] != order_df['trade_volume']
    order_df_cancel = order_df[order_df["canceled"]]
    # hprint(order_df)

    order_freq_df = pd.DataFrame(
        [
            {
                "production": p,
                "exch": exch,
                'direction': direct,
                "status": t0_status,
                "order_cnt": len(flag_df),
            } for (p, exch, direct, t0_status), flag_df in
            order_df.groupby(["production", "exch", "direction", "t0-status"])
        ]).set_index(["production", "exch", "direction", "status"])

    order_freq_df_cancel = pd.DataFrame(
        [
            {
                "production": p,
                "exch": exch,
                'direction': direct,
                "status": t0_status,
                "cancel_cnt": len(flag_df),
            } for (p, exch, direct, t0_status), flag_df in
            order_df_cancel.groupby(["production", "exch", "direction", "t0-status"])
        ]).set_index(["production", "exch", "direction", "status"])

    order_freq_df = pd.concat([
        order_freq_df, order_freq_df_cancel], axis=1).fillna(0).reset_index()
    order_freq_df['order_cnt'] = order_freq_df['order_cnt'].astype('int')
    order_freq_df['cancel_cnt'] = order_freq_df['cancel_cnt'].astype('int')
    order_freq_df.to_csv(f'{output_dir}{date}_order_num_details.csv', index=False)
    hprint(order_freq_df)


def plot_order_and_cancel_infor_daily(start_date, end_date, product):
    date_list = get_trading_days(start_date, end_date)
    conlist = []
    for date in date_list:
        print(date)
        df_infor = pd.read_csv(f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{date}/BrokerFeeSummary/'
                               f'{date}_daily_except_monitor_all.csv', encoding='GBK')

        df_infor = df_infor[df_infor['产品名'] == product]
        df_infor = df_infor[['exch', '总单数', '撤单率', '市值', '单边换手率']]
        df_infor['Date'] = date[2:]
        conlist.append(df_infor)

    df_infor = pd.concat(conlist, axis=0)
    df_orders = pd.pivot_table(df_infor, index='Date', columns='exch', values='总单数').rename(
        {'SZ': 'SZ-总单数', 'SH': 'SH-总单数'}, axis='columns')
    df_orders.index.name = None
    df_infor = df_infor.groupby('Date').agg(
        {'撤单率': 'mean', '市值': 'mean', '单边换手率': 'mean'}).rename({'市值': '市值(亿)'}, axis='columns')
    df_infor.index.name = None

    plt.figure(figsize=(20, 16))
    ax = plt.subplot(2, 2, 1)
    df_orders.plot.bar(ax=ax)
    format_ax(ax=ax)

    for col_i, col in enumerate(df_infor.columns.to_list()):
        ax = plt.subplot(2, 2, 2 + col_i)
        df_infor[[col]].plot.bar(ax=ax)
        format_ax(ax)

    plt.suptitle(f'{start_date}-{end_date} | {product}')
    plt.tight_layout()
    # plt.savefig(f'{LOG_TEMP_PATH}{flag}_自营.png')
    plt.show()


def generate_freq_report(curdate, wechat=False, output_dir=None):
    if output_dir is None:
        output_dir = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{curdate}/BrokerFeeSummary/'

    monitor_dir = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/OrderActionMonitor/'
    df_freq = pd.read_csv(f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{curdate}/OrderFreq/'
                          f'{curdate}_order_freq_1s.csv')
    df_accsumm = get_production_list_trading(curdate, ret_df_data=True)

    for paras in [
        'order_freq',
        'cancel_freq',
        'order_and_cancel_freq',
        'order_freq_shift',
        'cancel_freq_shift',
        'order_and_cancel_freq_shift',
        'order_and_cancel_freq_new',
        'order_daily_sum',
        'cancel_daily_sum',
        'order_and_cancel_daily_sum',
        'cancel_num_in_1s',
    ]:
        df_freq[paras] = df_freq[paras].astype('int')
    df_freq['cancel_freq'] = df_freq['cancel_freq'].astype('int')
    df_freq['order_and_cancel_freq_new'] = df_freq['order_and_cancel_freq_new'].astype('int')
    df_freq['cancel_num_in_1s'] = df_freq['cancel_num_in_1s'].astype('int')
    df_freq['cancel_daily_sum'] = df_freq['cancel_daily_sum'].astype('int')
    df_freq['order_and_cancel_daily_sum'] = df_freq['order_and_cancel_daily_sum'].astype('int')

    df_freq['cancel_diff_time_min'] = np.round(df_freq['cancel_diff_time_min'], 5)
    df_freq['cancel_diff_time_max'] = np.round(df_freq['cancel_diff_time_max'], 5)
    df_freq['cancel_diff_time_mean'] = np.round(df_freq['cancel_diff_time_mean'], 5)
    df_freq['cancel_diff_time_median'] = np.round(df_freq['cancel_diff_time_median'], 5)
    df_freq['cancel_ratio'] = np.round(df_freq['cancel_ratio'] * 100, 1)
    df_freq['市值'] = np.round(df_freq['production'].apply(lambda x: get_production_list_trading(curdate, x, ret_type='holdmv', df_accsum=df_accsumm)) / 1e8, 2)

    predate = get_predate(curdate, 1)
    conlist = []
    for prod in df_freq['production'].unique():
        df_position = get_position(predate, prod)
        df_position['production'] = prod
        conlist.append(df_position)

    df_position = pd.concat(conlist, axis=0)
    df_price = get_price(predate, predate).reset_index()[['SecuCode', 'ClosePrice']]
    df_position = pd.merge(df_position, df_price, on='SecuCode', how='left')
    df_position['市值(分沪深)'] = df_position['ClosePrice'] * df_position['Volume']
    df_position = df_position.groupby(['production', 'Exchange'])['市值(分沪深)'].sum().reset_index().rename(
        {"Exchange": 'exch'}, axis='columns')
    df_position['exch'] = df_position['exch'].replace({'SZE': 'SZ', 'SSE': 'SH'})

    df_freq = pd.merge(df_freq, df_position, on=['production', 'exch'], how='left')
    df_freq['单边换手率(分沪深)'] = np.round(df_freq['turnover'] / df_freq['市值(分沪深)'] / 2 * 100, 2)
    df_freq['turnover'] = np.round(df_freq['turnover'] / 1e8, 4)
    df_freq['市值(分沪深)'] = np.round(df_freq['市值(分沪深)'] / 1e8, 2)

    df_feesumm = pd.read_excel(f'{output_dir}{curdate}_daily_profit.xlsx')[['产品名', '策略', '单边换手率']]
    df_freq = df_freq.rename({
        'production': '产品名',
        'order_time_max': 'snd_t_max',
        'cancel_time_max': 'cncl_t_max',
        'order_and_cancel_time_max': 't_max',
        'order_freq': '发单速率',
        'cancel_freq': '撤单速率',
        'order_and_cancel_freq': '总速率',
        'order_freq_shift': '发单速率_shift',
        'cancel_freq_shift': '撤单速率_shift',
        'order_and_cancel_freq_shift': '总速率_shift',
        'order_and_cancel_freq_new': '总速率_new',
        'order_daily_sum': '发单数',
        'cancel_daily_sum': '撤单数',
        'order_and_cancel_daily_sum': '总单数',
        'cancel_num_in_1s': '秒内撤单数',
        'cancel_diff_time_min': '撤单时min',
        'cancel_diff_time_max': '撤单时max',
        'cancel_diff_time_mean': '撤单时(均)',
        'cancel_diff_time_median': '撤单时(中)',
        'cancel_ratio': '撤单率',
        'TpCode': '撤单TpCL',
        'CrTpCode': '撤单TpCrL',
        'turnover': '交易额(亿)',
    }, axis='columns')

    df_freq = pd.merge(df_freq, df_feesumm, on='产品名', how='left')
    df_freq = df_freq.sort_values(['产品名', 'exch'], ascending=True)
    # print(df_freq.head(10))
    df_freq.to_csv(f'{output_dir}{curdate}_daily_except_monitor_all.csv', index=False, encoding='GBK')

    df_freq_except_own = df_freq[df_freq['产品名'].isin(Production_OwnList)].reset_index(drop=True)
    df_freq_except_manage = df_freq[
        ((df_freq['总速率'] >= 300) |
        (df_freq['总单数'] >= 20000) |
        (df_freq['秒内撤单数'] > 100) |
        (df_freq['撤单率'] > 30) |
        (df_freq['撤单TpCrL'].apply(lambda x: float(x.split(',')[0])) > 5)
        ) &
        (~ df_freq['产品名'].isin(Production_OwnList))].reset_index(drop=True)

    freq_except_dict = {
        '_own': df_freq_except_own,
        '_manage': df_freq_except_manage
    }
    for freq_except in freq_except_dict:
        df_freq_except = freq_except_dict[freq_except]

        caption_html = f"""
                    <br><h3 style="margin: 0;">{f'{curdate}-当日交易参数监控{freq_except}'}</h3><br>
                    """
        df_freq_except = df_freq_except.astype('str').style.set_properties(**{'text-align': 'center'}).set_table_styles([
            {'selector': 'th', 'props': [('border', '1px solid black'), ('font-size', '10')]},
            {'selector': 'td', 'props': [('border', '1px solid black'), ('font-size', '10')]},
            {'selector': 'caption', 'props': [('border', '1px solid black'), ('font-size', '12')]}]).set_caption(
            caption_html).apply(
            lambda x: ['background-color: red' if float(_) > 300 else 'background-color: yellow' for _ in x], axis=0,
            subset=['总速率']).apply(
            lambda x: ['background-color: red' if float(_) > 2e4 else 'background-color: yellow' for _ in x], axis=0,
            subset=['总单数']).apply(
            lambda x: ['background-color: red' if float(_) > 400 else 'background-color: yellow' for _ in x], axis=0,
            subset=['秒内撤单数']).apply(
            lambda x: ['background-color: red' if float(_) > 30 else 'background-color: yellow' for _ in x], axis=0,
            subset=['撤单率']).apply(
            lambda x: ['background-color: green' if _ == 'SH' else '' for _ in x], axis=0, subset=['exch']).apply(
            lambda x: ['background-color: green'
                       if str(v) in Production_OwnList else '' for v in x], axis=0, subset=['产品名'])

        img_path = f'{output_dir}{curdate}_daily_except{freq_except}.png'
        try:
            dfi.export(df_freq_except, filename=img_path, dpi=300, max_cols=-1, max_rows=-1, table_conversion='selenium')
            shutil.copy(img_path, f'{monitor_dir}{curdate}_daily_except{freq_except}.png')
            if wechat:
                wechat_bot_image(img_path, type_api='analysis-report')
        except:
            pass

        # print(df_freq_except)


def generate_freq_report_pdf_monthly(end_date, start_date=None, wechat=False):
    month_date_list = get_monthly_trading_day(end_date)
    if start_date is None:
        if end_date == month_date_list[-1]:
            start_date = month_date_list[0]
            date_list = month_date_list
        else:
            start_date = get_predate(end_date, 20)
            date_list = get_trading_days(start_date, end_date)
    else:
        date_list = get_trading_days(start_date, end_date)

    output_dir = f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/%s/BrokerFeeSummary/'
    conlist = []
    for date in date_list:
        df_freq = pd.read_csv(f'{output_dir % date}{date}_daily_except_monitor_all.csv', encoding='GBK')
        df_freq['Date'] = date
        conlist.append(df_freq)
    df_freq = pd.concat(conlist, axis=0)

    summary_list = []
    for (prod, exch), df in df_freq.groupby(['产品名', 'exch']):
        df = df.set_index('Date')
        summary_list.append({
            '产品': prod,
            'exch': exch,
            '总速率(均)': int(round(df['总速率'].mean())),
            '总单数(均)': int(round(df['总单数'].mean())),
            '撤单率(均)': int(round(df['撤单率'].mean())),
            '秒内撤单数(均)': int(round(df['秒内撤单数'].mean())),
            '单边换手率(均)': int(round(df['撤单率'].mean())),

            '总速率(max)': df['总速率'].max(),
            '总速率(max)_date': df['总速率'].idxmax(),
            '总单数(max)': df['总单数'].max(),
            '总单数(max)_date': df['总单数'].idxmax(),
            '撤单率(max)': df['撤单率'].max(),
            '撤单率(max)_date': df['撤单率'].idxmax(),
            '秒内撤单数(max)': df['秒内撤单数'].max(),
            '秒内撤单数(max)_date': df['秒内撤单数'].idxmax(),
            '单边换手率(max)': df['撤单率'].max(),
            '单边换手率(max)_date': df['撤单率'].idxmax(),

            '市值': round(df['市值'].mean(), 2),
            '策略': ",".join(df['策略'].dropna(axis=0).astype('str').unique()),
        })

    df_freq = pd.DataFrame(summary_list).sort_values(['exch', '产品'])
    df_freq.to_csv(
        f'{output_dir % end_date}{start_date}_{end_date}_daily_except_monitor_all.csv', index=False, encoding='GBK')

    df_freq_except = df_freq[
        (df_freq['总速率(max)'] >= 300) |
        (df_freq['总单数(max)'] >= 20000) |
        (df_freq['秒内撤单数(max)'] > 100) |
        (df_freq['撤单率(max)'] > 30)].reset_index(drop=True)
    df_freq_except = df_freq_except[
        df_freq_except['产品'].isin(WinterFallProductionList)].reset_index(drop=True)

    for exchange in ['SH', 'SZ']:
        df_freq_except_exch = df_freq_except[df_freq_except['exch'] == exchange]
        len_df = len(df_freq_except_exch)
        caption_html = f"""
                        <br><h3 style="margin: 0;">{f'{start_date}-{end_date}[{exchange}]交易参数监控'}</h3><br>
                        """
        df_freq_except_exch = df_freq_except_exch.astype('str').style.set_properties(
            **{'text-align': 'center'}).set_table_styles([
            {'selector': 'th', 'props': [('border', '1px solid black'), ('font-size', '10')]},
            {'selector': 'td', 'props': [('border', '1px solid black'), ('font-size', '10')]},
            {'selector': 'caption', 'props': [('border', '1px solid black'), ('font-size', '12')]}]).set_caption(
            caption_html).apply(
            lambda x: ['background-color: red' if float(_) > 300 else 'background-color: yellow' for _ in x], axis=0,
            subset=['总速率(max)', '总速率(均)']).apply(
            lambda x: ['background-color: red' if float(_) > 2e4 else 'background-color: yellow' for _ in x], axis=0,
            subset=['总单数(max)', '总单数(均)']).apply(
            lambda x: ['background-color: red' if float(_) > 400 else 'background-color: yellow' for _ in x], axis=0,
            subset=['秒内撤单数(max)', '秒内撤单数(均)']).apply(
            lambda x: ['background-color: red' if float(_) > 30 else 'background-color: yellow' for _ in x], axis=0,
            subset=['撤单率(max)', '撤单率(均)']).apply(
        lambda x: ['background-color: green'
                   if str(v) in Production_OwnList else '' for v in x], axis=0, subset=['产品'])

        if len_df < 100:
            img_path = f'{output_dir % end_date}{start_date}_{end_date}_longtime_except_{exchange}.png'
            dfi.export(df_freq_except_exch, filename=img_path, dpi=300, max_cols=-1, max_rows=-1, table_conversion='chrome')
        else:
            file_path = f'{output_dir % end_date}{start_date}_{end_date}_longtime_except_{exchange}.xlsx'
            df_freq_except_exch.to_excel(file_path, index=False, encoding='GBK')


if __name__ == "__main__":
    curdate = datetime.datetime.now().strftime('%Y%m%d')
    # curdate = '20231115'
    # if len(sys.argv) < 2:
    #     op_date = curdate
    # else:
    #     op_date = sys.argv[1]
    #
    # freq_analysis(op_date)

    order_freq_analysis_monthly('20241008', '20241031')

    # pool_starmap_multiprocessing(
    #     freq_analysis,
    #     [
    #         [
    #             date,
    #             f'{PLATFORM_PATH_DICT["v_path"]}StockData/d0_file/IntraAnalysisResults/{date}/OrderFreq/'
    #         ]
    #         for date in get_trading_days('20230601', '20231103')
    #     ],
    #     30)

    # for op_date in get_trading_days('20230601', '20231103')[::-1]:
    #     freq_analysis(op_date)
    #     freq_analysis_zzzq1(op_date)

    # load_alpha_order_md(op_date, "DC9")
